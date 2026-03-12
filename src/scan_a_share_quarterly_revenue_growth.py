from __future__ import annotations

import argparse
import json
import math
import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


QUARTER_LABEL_MAP = {
    "Q1": "第一季度",
    "Q2": "第二季度",
    "Q3": "第三季度",
    "Q4": "第四季度",
}

QUARTER_ORDER = ["Q1", "Q2", "Q3", "Q4"]

RESULT_HEADERS = [
    "股票代码",
    "证券代码",
    "股票简称",
    "所属市场",
    "所属三级行业",
    "年份",
    "季度",
    "报告类型",
    "季度报告期",
    "季度报发布时间",
    "公司总市值范围",
    "上市日期",
    "单季度营收同比增速_百分比",
]


class RetryClient:
    def __init__(self, max_retry: int, timeout_sec: int = 45) -> None:
        self.max_retry = max_retry
        self.timeout_sec = timeout_sec
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            }
        )

    def get_json(self, url: str, timeout_sec: int | None = None) -> dict[str, Any]:
        timeout = timeout_sec or self.timeout_sec
        last_exc: Exception | None = None
        for attempt in range(1, self.max_retry + 1):
            try:
                resp = self.session.get(url, timeout=timeout)
                resp.raise_for_status()
                return resp.json()
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if attempt >= self.max_retry:
                    break
                time.sleep(2 ** attempt)
        raise RuntimeError(f"Request failed after retries: {url}") from last_exc

    def get_text(self, url: str, timeout_sec: int | None = None) -> str:
        timeout = timeout_sec or self.timeout_sec
        last_exc: Exception | None = None
        for attempt in range(1, self.max_retry + 1):
            try:
                resp = self.session.get(url, timeout=timeout)
                resp.raise_for_status()
                resp.encoding = resp.apparent_encoding or resp.encoding
                return resp.text
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if attempt >= self.max_retry:
                    break
                time.sleep(2 ** attempt)
        raise RuntimeError(f"Request failed after retries: {url}") from last_exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="扫描A股单季度营收同比增速")
    parser.add_argument("--growth-threshold", type=float, default=20.0, help="同比增速阈值，默认20")
    parser.add_argument("--year", type=int, default=0, help="目标年份，例如 2025")
    parser.add_argument("--quarter", choices=QUARTER_ORDER, default="", help="目标季度：Q1/Q2/Q3/Q4")
    parser.add_argument("--output-path", default="", help="输出 Excel 文件路径")
    parser.add_argument("--max-retry", type=int, default=4, help="接口重试次数，默认4")
    return parser.parse_args()


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None

    raw = str(value).strip()
    if not raw or raw in {"--", "-", "None", "null"}:
        return None

    if re.fullmatch(r"\d{8}", raw):
        try:
            return datetime.strptime(raw, "%Y%m%d")
        except ValueError:
            return None

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue

    if "T" in raw:
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            pass

    if len(raw) >= 10 and re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", raw):
        try:
            return datetime.strptime(raw[:10], "%Y-%m-%d")
        except ValueError:
            return None

    return None


def parse_positive_float(value: Any) -> float | None:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    if num <= 0:
        return None
    return num


def parse_numeric(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def get_quarter_from_month(month: int) -> str | None:
    return {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(month)


def get_single_quarter_revenue(cum_by_quarter: dict[str, float], quarter: str) -> float | None:
    if quarter == "Q1":
        return cum_by_quarter.get("Q1")
    if quarter == "Q2" and "Q2" in cum_by_quarter and "Q1" in cum_by_quarter:
        return cum_by_quarter["Q2"] - cum_by_quarter["Q1"]
    if quarter == "Q3" and "Q3" in cum_by_quarter and "Q2" in cum_by_quarter:
        return cum_by_quarter["Q3"] - cum_by_quarter["Q2"]
    if quarter == "Q4" and "Q4" in cum_by_quarter and "Q3" in cum_by_quarter:
        return cum_by_quarter["Q4"] - cum_by_quarter["Q3"]
    return None


def get_market_cap_range(market_cap: float | None) -> str:
    if market_cap is None or market_cap <= 0:
        return ""
    if market_cap < 10_000_000_000:
        return "0-100亿"
    if market_cap < 50_000_000_000:
        return "100-500亿"
    if market_cap < 100_000_000_000:
        return "500-1000亿"
    return "1000亿以上"


def get_financial_page_data(client: RetryClient, encoded_filter: str, page_number: int) -> dict[str, Any]:
    base_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
    query = (
        "reportName=RPT_LICO_FN_CPD"
        "&columns=SECURITY_CODE,SECUCODE,SECURITY_NAME_ABBR,SECURITY_TYPE,"
        "TRADE_MARKET,DATATYPE,REPORTDATE,NOTICE_DATE,TOTAL_OPERATE_INCOME"
        "&source=WEB&client=WEB&pageSize=500"
        f"&filter={encoded_filter}&pageNumber={page_number}"
    )
    url = f"{base_url}?{query}"
    resp = client.get_json(url)
    result = resp.get("result") or {}
    data = result.get("data") or []
    if not isinstance(data, list):
        data = []
    return {
        "Data": data,
        "Pages": int(result.get("pages") or 0),
        "Count": int(result.get("count") or 0),
    }


def get_market_snapshot_map(client: RetryClient) -> dict[str, dict[str, Any]]:
    base_url = "https://push2.eastmoney.com/api/qt/clist/get"
    page_size = 500
    common_query = (
        f"pz={page_size}&po=1&np=1"
        "&ut=bd1d9ddb04089700cf9c27f6f7426281"
        "&fltt=2&invt=2&fid=f3"
        "&fs=m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23"
        "&fields=f12,f20,f26,f100"
    )

    first_resp = client.get_json(f"{base_url}?pn=1&{common_query}")
    first_data = first_resp.get("data") or {}
    total = int(first_data.get("total") or 0)
    pages = int(math.ceil(total / page_size)) if total > 0 else 0
    pages = max(pages, 1)

    snapshot_map: dict[str, dict[str, Any]] = {}

    def add_diff(diff: list[dict[str, Any]] | None) -> None:
        if not diff:
            return
        for row in diff:
            code = str(row.get("f12") or "").strip()
            if not code:
                continue

            market_cap = parse_positive_float(row.get("f20"))
            listing_date = parse_datetime(row.get("f26"))
            industry_name = str(row.get("f100") or "")

            snapshot_map[code] = {
                "TotalMarketCap": market_cap,
                "ListingDate": listing_date,
                "IndustryName": industry_name,
            }

    add_diff(first_data.get("diff"))
    for page in range(2, pages + 1):
        resp = client.get_json(f"{base_url}?pn={page}&{common_query}")
        add_diff((resp.get("data") or {}).get("diff"))

    return snapshot_map


def get_secid(security_code: str, secucode: str) -> str:
    upper = secucode.upper()
    if upper.endswith(".SH"):
        return f"1.{security_code}"
    if upper.endswith(".SZ"):
        return f"0.{security_code}"
    if re.match(r"^(60|68|90)", security_code):
        return f"1.{security_code}"
    return f"0.{security_code}"


def get_stock_detail_by_secid(client: RetryClient, secid: str) -> dict[str, Any]:
    url = (
        "https://push2.eastmoney.com/api/qt/stock/get"
        f"?secid={secid}&ut=fa5fd1943c7b386f172d6893dbfba10b"
        "&fields=f57,f116,f189,f127,f100"
    )
    resp = client.get_json(url)
    data = resp.get("data") or {}
    industry = str(data.get("f127") or "").strip() or str(data.get("f100") or "").strip()
    return {
        "TotalMarketCap": parse_positive_float(data.get("f116")),
        "ListingDate": parse_datetime(data.get("f189")),
        "IndustryL3": industry,
    }


def get_listing_date_from_sina(client: RetryClient, security_code: str) -> datetime | None:
    url = f"https://vip.stock.finance.sina.com.cn/corp/go.php/vCI_CorpInfo/stockid/{security_code}.phtml"
    try:
        content = client.get_text(url, timeout_sec=30)
    except Exception:  # noqa: BLE001
        return None

    for pattern, fmt in (
        (r"上市日期：</td>\s*<td class=\"cc\">\s*([0-9]{4}-[0-9]{2}-[0-9]{2})", "%Y-%m-%d"),
        (r"上市日期：</td>\s*<td class=\"cc\">\s*([0-9]{4}/[0-9]{2}/[0-9]{2})", "%Y/%m/%d"),
    ):
        match = re.search(pattern, content)
        if not match:
            continue
        try:
            return datetime.strptime(match.group(1), fmt)
        except ValueError:
            continue
    return None


def ask_for_year() -> int:
    year_text = input("请输入目标年份(例如 2025): ").strip()
    if not year_text.isdigit():
        raise ValueError("年份输入无效，请输入四位数字年份，例如 2025。")
    return int(year_text)


def ask_for_quarter() -> str:
    print("请选择单季度类型：")
    print("1. Q1（第一季度）")
    print("2. Q2（第二季度）")
    print("3. Q3（第三季度）")
    print("4. Q4（第四季度）")
    choice = input("请输入编号(1/2/3/4): ").strip()
    mapping = {"1": "Q1", "2": "Q2", "3": "Q3", "4": "Q4"}
    if choice not in mapping:
        raise ValueError("无效输入。请重新运行并输入 1/2/3/4。")
    return mapping[choice]


def normalize_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in raw_rows:
        if row.get("SECURITY_TYPE") != "A股":
            continue

        income = parse_numeric(row.get("TOTAL_OPERATE_INCOME"))
        if income is None:
            continue

        report_date = parse_datetime(row.get("REPORTDATE"))
        if report_date is None:
            continue

        quarter = get_quarter_from_month(report_date.month)
        if quarter is None:
            continue

        notice_date = parse_datetime(row.get("NOTICE_DATE")) or report_date
        normalized.append(
            {
                "SecurityCode": str(row.get("SECURITY_CODE") or ""),
                "SecuCode": str(row.get("SECUCODE") or ""),
                "Name": str(row.get("SECURITY_NAME_ABBR") or ""),
                "TradeMarket": str(row.get("TRADE_MARKET") or ""),
                "ReportType": str(row.get("DATATYPE") or ""),
                "ReportDate": report_date,
                "NoticeDate": notice_date,
                "Year": report_date.year,
                "Quarter": quarter,
                "CumRevenue": income,
            }
        )
    return normalized


def pick_latest_cum_rows(normalized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest_map: dict[tuple[str, int, str], dict[str, Any]] = {}
    for row in normalized_rows:
        key = (row["SecurityCode"], row["Year"], row["Quarter"])
        old = latest_map.get(key)
        if old is None or row["NoticeDate"] > old["NoticeDate"]:
            latest_map[key] = row
    return list(latest_map.values())


def build_single_quarter_rows(latest_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in latest_rows:
        grouped[(row["SecurityCode"], row["Year"])].append(row)

    output: list[dict[str, Any]] = []
    for items in grouped.values():
        cum_by_quarter: dict[str, float] = {}
        row_by_quarter: dict[str, dict[str, Any]] = {}
        for item in items:
            cum_by_quarter[item["Quarter"]] = item["CumRevenue"]
            row_by_quarter[item["Quarter"]] = item

        for quarter in QUARTER_ORDER:
            single = get_single_quarter_revenue(cum_by_quarter, quarter)
            if single is None or quarter not in row_by_quarter:
                continue
            q_row = row_by_quarter[quarter]
            output.append(
                {
                    "SecurityCode": q_row["SecurityCode"],
                    "SecuCode": q_row["SecuCode"],
                    "Name": q_row["Name"],
                    "TradeMarket": q_row["TradeMarket"],
                    "Year": int(q_row["Year"]),
                    "Quarter": quarter,
                    "QuarterReportType": q_row["ReportType"],
                    "QuarterReportDate": q_row["ReportDate"],
                    "QuarterNoticeDate": q_row["NoticeDate"],
                    "SingleRevenue": float(single),
                }
            )
    return output


def auto_fit_worksheet(worksheet: Any) -> None:
    for idx, col in enumerate(worksheet.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)


def write_sheet(worksheet: Any, rows: list[dict[str, Any]], headers: list[str] | None = None) -> None:
    actual_headers = headers or (list(rows[0].keys()) if rows else [])
    if actual_headers:
        worksheet.append(actual_headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in actual_headers])
    auto_fit_worksheet(worksheet)


def add_medium6_table(worksheet: Any, table_name: str) -> None:
    if worksheet.max_column <= 0 or worksheet.max_row <= 0:
        return

    # Excel table requires a rectangular range with at least header row.
    max_col_letter = get_column_letter(worksheet.max_column)
    table_ref = f"A1:{max_col_letter}{worksheet.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_values_keep_style(worksheet: Any, rows: list[dict[str, Any]], headers: list[str]) -> None:
    max_cols = len(headers)
    target_last_row = max(2, len(rows) + 1)
    clear_to_row = max(worksheet.max_row, target_last_row)

    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)

    for row_idx in range(2, clear_to_row + 1):
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    # Ensure table has at least one data row, even when result is empty.
    if not rows:
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=2, column=col_idx, value="")


def upsert_table_ref(
    worksheet: Any,
    table_name: str,
    max_cols: int,
    data_rows: int,
    default_style_name: str = "TableStyleMedium6",
) -> None:
    table_last_row = max(2, data_rows + 1)
    table_ref = f"A1:{get_column_letter(max_cols)}{table_last_row}"

    if table_name in worksheet.tables:
        table = worksheet.tables[table_name]
        table.ref = table_ref
        if table.tableStyleInfo is None:
            table.tableStyleInfo = TableStyleInfo(
                name=default_style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        return

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=default_style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def build_workbook_with_template(
    template_path: Path,
    summary: list[dict[str, Any]],
    result: list[dict[str, Any]],
) -> Workbook:
    workbook = load_workbook(template_path)
    if "Summary" not in workbook.sheetnames or "Result" not in workbook.sheetnames:
        raise ValueError("样式模板必须包含 Summary 和 Result 两个工作表。")

    ws_summary = workbook["Summary"]
    ws_result = workbook["Result"]

    summary_headers = ["指标", "值"]
    write_values_keep_style(ws_summary, summary, summary_headers)
    upsert_table_ref(ws_summary, "SummaryTable", len(summary_headers), len(summary))

    write_values_keep_style(ws_result, result, RESULT_HEADERS)
    upsert_table_ref(ws_result, "SingleQuarterRevenueYoYResult", len(RESULT_HEADERS), len(result))

    return workbook


def main() -> None:
    args = parse_args()

    year = args.year if args.year > 0 else ask_for_year()
    max_allowed_year = datetime.now().year + 1
    if year < 2008 or year > max_allowed_year:
        raise ValueError(f"年份超出允许范围。当前支持 2008 ~ {max_allowed_year}。")

    quarter = args.quarter or ask_for_quarter()
    growth_threshold = args.growth_threshold

    if args.output_path:
        output_path = Path(args.output_path)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path.cwd() / "output" / (
            f"a_share_{year}_{quarter}_single_revenue_yoy_gt{int(growth_threshold)}_{stamp}.xlsx"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    start_date = f"{year - 1}-01-01"
    end_date = f"{year}-12-31"
    encoded_filter = quote(f"(REPORTDATE>='{start_date}')(REPORTDATE<='{end_date}')", safe="")

    client = RetryClient(max_retry=args.max_retry, timeout_sec=45)
    first_page = get_financial_page_data(client, encoded_filter, 1)

    all_rows: list[dict[str, Any]] = list(first_page["Data"])
    total_pages = first_page["Pages"]
    for page_number in range(2, total_pages + 1):
        page_data = get_financial_page_data(client, encoded_filter, page_number)
        all_rows.extend(page_data["Data"])

    normalized = normalize_rows(all_rows)
    latest_cum_rows = pick_latest_cum_rows(normalized)
    single_quarter_rows = build_single_quarter_rows(latest_cum_rows)

    target_rows = [r for r in single_quarter_rows if r["Year"] == year and r["Quarter"] == quarter]
    prev_rows = [r for r in single_quarter_rows if r["Year"] == year - 1 and r["Quarter"] == quarter]
    prev_map = {r["SecurityCode"]: r for r in prev_rows}

    try:
        market_map = get_market_snapshot_map(client)
    except Exception:  # noqa: BLE001
        market_map = {}
    stock_detail_cache: dict[str, dict[str, Any]] = {}
    sina_listing_cache: dict[str, datetime | None] = {}
    supplement_hit_count = 0

    merged: list[dict[str, Any]] = []
    for cur in target_rows:
        prev = prev_map.get(cur["SecurityCode"])
        if prev is None:
            continue
        prev_revenue = prev.get("SingleRevenue")
        if prev_revenue in (None, 0):
            continue

        yoy = ((cur["SingleRevenue"] - prev_revenue) / prev_revenue) * 100

        market_cap: float | None = None
        listing_date_text = ""
        industry_l3 = ""

        snap = market_map.get(cur["SecurityCode"])
        if snap:
            market_cap = snap.get("TotalMarketCap")
            if snap.get("ListingDate"):
                listing_date_text = snap["ListingDate"].strftime("%Y-%m-%d")
            industry_l3 = str(snap.get("IndustryName") or "")

        need_supplement = market_cap is None or not listing_date_text or not industry_l3
        if need_supplement:
            secid = get_secid(cur["SecurityCode"], cur["SecuCode"])
            detail = stock_detail_cache.get(secid)
            if detail is None:
                try:
                    detail = get_stock_detail_by_secid(client, secid)
                except Exception:  # noqa: BLE001
                    detail = {"TotalMarketCap": None, "ListingDate": None, "IndustryL3": ""}
                stock_detail_cache[secid] = detail

            supplement_used = False
            if market_cap is None and detail.get("TotalMarketCap") is not None:
                market_cap = detail["TotalMarketCap"]
                supplement_used = True
            if not listing_date_text and detail.get("ListingDate") is not None:
                listing_date_text = detail["ListingDate"].strftime("%Y-%m-%d")
                supplement_used = True
            if not industry_l3 and detail.get("IndustryL3"):
                industry_l3 = detail["IndustryL3"]
                supplement_used = True

            if not listing_date_text:
                sina_date = sina_listing_cache.get(cur["SecurityCode"])
                if cur["SecurityCode"] not in sina_listing_cache:
                    sina_date = get_listing_date_from_sina(client, cur["SecurityCode"])
                    sina_listing_cache[cur["SecurityCode"]] = sina_date
                if sina_date:
                    listing_date_text = sina_date.strftime("%Y-%m-%d")
                    supplement_used = True

            if supplement_used:
                supplement_hit_count += 1

        merged.append(
            {
                "股票代码": cur["SecurityCode"],
                "证券代码": cur["SecuCode"],
                "股票简称": cur["Name"],
                "所属市场": cur["TradeMarket"],
                "所属三级行业": industry_l3,
                "年份": year,
                "季度": quarter,
                "报告类型": cur["QuarterReportType"],
                "季度报告期": cur["QuarterReportDate"].strftime("%Y-%m-%d"),
                "季度报发布时间": cur["QuarterNoticeDate"].strftime("%Y-%m-%d"),
                "公司总市值范围": get_market_cap_range(market_cap),
                "上市日期": listing_date_text,
                "单季度营收同比增速_百分比": round(yoy, 2),
            }
        )

    result = sorted(
        [row for row in merged if row["单季度营收同比增速_百分比"] > growth_threshold],
        key=lambda row: row["单季度营收同比增速_百分比"],
        reverse=True,
    )

    summary = [
        {"指标": "执行时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"指标": "目标年份", "值": year},
        {"指标": "目标季度", "值": f"{quarter}({QUARTER_LABEL_MAP[quarter]})"},
        {"指标": "扫描区间", "值": f"{start_date} ~ {end_date}"},
        {"指标": "接口总记录数", "值": first_page["Count"]},
        {"指标": "A股累计营收记录数", "值": len(normalized)},
        {"指标": "可计算目标季度同比公司数", "值": len(merged)},
        {"指标": "补充公开数据命中数", "值": supplement_hit_count},
        {"指标": "筛选条件", "值": f"单季度营收同比增速 > {growth_threshold}%"},
        {"指标": "命中公司数", "值": len(result)},
    ]

    workbook: Workbook | None = None
    repo_root = Path(__file__).resolve().parent.parent
    template_path = repo_root / "docs" / "templates" / "output_template.xlsx"
    if template_path.exists():
        try:
            workbook = build_workbook_with_template(template_path, summary, result)
        except Exception:  # noqa: BLE001
            workbook = None

    if workbook is None:
        workbook = Workbook()
        ws_summary = workbook.active
        ws_summary.title = "Summary"
        write_sheet(ws_summary, summary, headers=["指标", "值"])
        add_medium6_table(ws_summary, "SummaryTable")

        ws_result = workbook.create_sheet("Result")
        write_sheet(ws_result, result, headers=RESULT_HEADERS)
        add_medium6_table(ws_result, "SingleQuarterRevenueYoYResult")
    workbook.save(output_path)

    payload = {
        "output_path": str(output_path.resolve()),
        "year": year,
        "quarter": quarter,
        "api_total_records": first_page["Count"],
        "a_share_records": len(normalized),
        "comparable_companies": len(merged),
        "supplemented_rows": supplement_hit_count,
        "matched_companies": len(result),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
