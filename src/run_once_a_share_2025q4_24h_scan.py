from __future__ import annotations

import argparse
import json
import math
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

RESULT_HEADERS = [
    "股票代码",
    "证券代码",
    "股票简称",
    "所属市场",
    "所属三级行业",
    "年份",
    "季度",
    "报告类型",
    "季度报告期",
    "季度报发布时间",
    "公司总市值范围",
    "上市日期",
    "单季度营收同比增速_百分比",
]


class RetryClient:
    def __init__(self, max_retry: int, timeout_sec: int = 45) -> None:
        self.max_retry = max_retry
        self.timeout_sec = timeout_sec
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            }
        )

    def get_json(self, url: str, timeout_sec: int | None = None) -> dict[str, Any]:
        timeout = timeout_sec or self.timeout_sec
        last_exc: Exception | None = None
        for attempt in range(1, self.max_retry + 1):
            try:
                response = self.session.get(url, timeout=timeout)
                response.raise_for_status()
                return response.json()
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if attempt >= self.max_retry:
                    break
                time.sleep(2**attempt)
        raise RuntimeError(f"Request failed after retries: {url}") from last_exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="A股单季度营收同比（24小时公告窗口）扫描")
    parser.add_argument("--target-year", "--TargetYear", dest="target_year", type=int, default=2025, help="目标年份，默认2025")
    parser.add_argument(
        "--target-quarter",
        "--TargetQuarter",
        dest="target_quarter",
        choices=QUARTERS,
        default="Q4",
        help="目标季度，默认Q4",
    )
    parser.add_argument(
        "--growth-threshold",
        "--GrowthThreshold",
        dest="growth_threshold",
        type=float,
        default=20.0,
        help="同比增速阈值，默认20",
    )
    parser.add_argument(
        "--notice-within-hours",
        "--NoticeWithinHours",
        dest="notice_within_hours",
        type=int,
        default=24,
        help="公告发布时间窗口（小时），默认24，范围1-168",
    )
    parser.add_argument(
        "--output-dir",
        "--OutputDir",
        dest="output_dir",
        default="",
        help=r"输出目录，默认 Analysis/output",
    )
    parser.add_argument("--max-retry", "--MaxRetry", dest="max_retry", type=int, default=4, help="接口重试次数，默认4")
    return parser.parse_args()


def parse_numeric(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw in {"--", "-", "None", "null", ""}:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def get_quarter_from_month(month: int) -> str | None:
    return {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(month)


def get_single_quarter_revenue(cum_map: dict[str, float], quarter: str) -> float | None:
    if quarter == "Q1":
        return cum_map.get("Q1")
    if quarter == "Q2" and "Q2" in cum_map and "Q1" in cum_map:
        return cum_map["Q2"] - cum_map["Q1"]
    if quarter == "Q3" and "Q3" in cum_map and "Q2" in cum_map:
        return cum_map["Q3"] - cum_map["Q2"]
    if quarter == "Q4" and "Q4" in cum_map and "Q3" in cum_map:
        return cum_map["Q4"] - cum_map["Q3"]
    return None


def get_market_cap_range(market_cap: float | None) -> str:
    if market_cap is None or market_cap < 0:
        return ""
    if market_cap < 10_000_000_000:
        return "0-100亿"
    if market_cap < 50_000_000_000:
        return "100-500亿"
    if market_cap < 100_000_000_000:
        return "500-1000亿"
    return "1000亿以上"


def get_financial_page_data(client: RetryClient, page_number: int, filter_expr: str) -> dict[str, Any]:
    base_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
    query = (
        "reportName=RPT_LICO_FN_CPD"
        "&columns=SECURITY_CODE,SECUCODE,SECURITY_NAME_ABBR,SECURITY_TYPE,TRADE_MARKET,"
        "DATATYPE,REPORTDATE,NOTICE_DATE,TOTAL_OPERATE_INCOME"
        "&source=WEB"
        "&client=WEB"
        f"&pageNumber={page_number}"
        "&pageSize=500"
        f"&filter={filter_expr}"
    )
    response = client.get_json(f"{base_url}?{query}")
    result = response.get("result") or {}
    data = result.get("data") or []
    if not isinstance(data, list):
        data = []
    return {
        "Data": data,
        "Pages": int(result.get("pages") or 0),
        "Count": int(result.get("count") or 0),
    }


def get_market_snapshot_map(client: RetryClient) -> dict[str, dict[str, Any]]:
    snapshot: dict[str, dict[str, Any]] = {}

    base_url = "https://push2.eastmoney.com/api/qt/clist/get"
    page_size = 500
    query = (
        f"pz={page_size}&po=1&np=1"
        "&ut=bd1d9ddb04089700cf9c27f6f7426281"
        "&fltt=2&invt=2&fid=f3"
        "&fs=m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23"
        "&fields=f12,f20,f26,f100"
    )

    def add_diff(diff: list[dict[str, Any]] | None) -> None:
        if not diff:
            return
        for row in diff:
            code = str(row.get("f12") or "").strip()
            if not code:
                continue

            market_cap = parse_numeric(row.get("f20"))
            if market_cap is not None and market_cap < 0:
                market_cap = None

            listing_date = None
            listing_raw = str(row.get("f26") or "").strip()
            if len(listing_raw) == 8 and listing_raw.isdigit():
                try:
                    listing_date = datetime.strptime(listing_raw, "%Y%m%d")
                except ValueError:
                    listing_date = None

            industry_l3 = str(row.get("f100") or "").strip()
            if industry_l3 == "-":
                industry_l3 = ""

            snapshot[code] = {
                "TotalMarketCap": market_cap,
                "ListingDate": listing_date,
                "IndustryL3": industry_l3,
            }

    first = client.get_json(f"{base_url}?pn=1&{query}")
    first_data = first.get("data") or {}
    total = int(first_data.get("total") or 0)
    pages = int(math.ceil(total / page_size)) if total > 0 else 0
    pages = max(pages, 1)

    add_diff(first_data.get("diff"))
    for page_number in range(2, pages + 1):
        resp = client.get_json(f"{base_url}?pn={page_number}&{query}")
        add_diff((resp.get("data") or {}).get("diff"))

    return snapshot


def normalize_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in raw_rows:
        if row.get("SECURITY_TYPE") != "A股":
            continue

        income = parse_numeric(row.get("TOTAL_OPERATE_INCOME"))
        if income is None:
            continue

        report_date = parse_datetime(row.get("REPORTDATE"))
        notice_date = parse_datetime(row.get("NOTICE_DATE"))
        if report_date is None or notice_date is None:
            continue

        quarter = get_quarter_from_month(report_date.month)
        if quarter is None:
            continue

        normalized.append(
            {
                "SecurityCode": str(row.get("SECURITY_CODE") or ""),
                "SecuCode": str(row.get("SECUCODE") or ""),
                "Name": str(row.get("SECURITY_NAME_ABBR") or ""),
                "TradeMarket": str(row.get("TRADE_MARKET") or ""),
                "ReportType": str(row.get("DATATYPE") or ""),
                "ReportDate": report_date,
                "NoticeDate": notice_date,
                "Year": report_date.year,
                "Quarter": quarter,
                "CumRevenue": income,
            }
        )
    return normalized


def pick_latest_cum_rows(normalized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[tuple[str, int, str], dict[str, Any]] = {}
    for row in normalized_rows:
        key = (row["SecurityCode"], row["Year"], row["Quarter"])
        old = latest.get(key)
        if old is None or row["NoticeDate"] > old["NoticeDate"]:
            latest[key] = row
    return list(latest.values())


def build_single_quarter_rows(latest_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in latest_rows:
        grouped[(row["SecurityCode"], row["Year"])].append(row)

    output: list[dict[str, Any]] = []
    for group_rows in grouped.values():
        cum_map: dict[str, float] = {}
        row_map: dict[str, dict[str, Any]] = {}
        for item in group_rows:
            quarter = item["Quarter"]
            cum_map[quarter] = item["CumRevenue"]
            row_map[quarter] = item

        for quarter in QUARTERS:
            single_revenue = get_single_quarter_revenue(cum_map, quarter)
            if single_revenue is None or quarter not in row_map:
                continue

            row = row_map[quarter]
            output.append(
                {
                    "SecurityCode": row["SecurityCode"],
                    "SecuCode": row["SecuCode"],
                    "Name": row["Name"],
                    "TradeMarket": row["TradeMarket"],
                    "Year": int(row["Year"]),
                    "Quarter": quarter,
                    "QuarterReportType": row["ReportType"],
                    "QuarterReportDate": row["ReportDate"],
                    "QuarterNoticeDate": row["NoticeDate"],
                    "SingleRevenue": float(single_revenue),
                }
            )
    return output


def auto_fit_worksheet(worksheet: Any) -> None:
    for idx, col in enumerate(worksheet.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)


def write_values_keep_style(worksheet: Any, rows: list[dict[str, Any]], headers: list[str]) -> None:
    max_cols = len(headers)
    target_last_row = max(2, len(rows) + 1)
    clear_to_row = max(worksheet.max_row, target_last_row)

    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)

    for row_idx in range(2, clear_to_row + 1):
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    if not rows:
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=2, column=col_idx, value="")

    auto_fit_worksheet(worksheet)


def upsert_table(
    worksheet: Any,
    table_name: str,
    max_cols: int,
    data_rows: int,
    style_name: str = "TableStyleMedium6",
) -> None:
    table_last_row = max(2, data_rows + 1)
    table_ref = f"A1:{get_column_letter(max_cols)}{table_last_row}"
    if table_name in worksheet.tables:
        table = worksheet.tables[table_name]
        table.ref = table_ref
        if table.tableStyleInfo is None:
            table.tableStyleInfo = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        return

    if worksheet.tables:
        existing = next(iter(worksheet.tables.values()))
        existing.ref = table_ref
        if existing.tableStyleInfo is None:
            existing.tableStyleInfo = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        return

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def build_workbook_with_template(
    template_path: Path,
    summary_rows: list[dict[str, Any]],
    result_rows: list[dict[str, Any]],
) -> Workbook:
    workbook = load_workbook(template_path)
    if "Summary" not in workbook.sheetnames or "Result" not in workbook.sheetnames:
        raise ValueError("样式模板必须包含 Summary 和 Result 两个工作表。")

    summary_sheet = workbook["Summary"]
    result_sheet = workbook["Result"]

    summary_headers = ["指标", "值"]
    write_values_keep_style(summary_sheet, summary_rows, summary_headers)
    write_values_keep_style(result_sheet, result_rows, RESULT_HEADERS)

    upsert_table(summary_sheet, "SummaryTable", len(summary_headers), len(summary_rows))
    upsert_table(result_sheet, "ResultTable", len(RESULT_HEADERS), len(result_rows))
    return workbook


def build_workbook_fallback(summary_rows: list[dict[str, Any]], result_rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_headers = ["指标", "值"]
    write_values_keep_style(summary_sheet, summary_rows, summary_headers)
    write_values_keep_style(workbook.create_sheet("Result"), result_rows, RESULT_HEADERS)

    upsert_table(summary_sheet, "SummaryTable", len(summary_headers), len(summary_rows))
    upsert_table(workbook["Result"], "ResultTable", len(RESULT_HEADERS), len(result_rows))
    return workbook


def main() -> None:
    args = parse_args()

    if not (1 <= args.notice_within_hours <= 168):
        raise ValueError("--notice-within-hours 必须在 1~168 之间。")

    run_time = datetime.now()
    notice_window_start = run_time - timedelta(hours=args.notice_within_hours)

    start_date = f"{args.target_year - 1}-01-01"
    end_date = f"{args.target_year}-12-31"
    filter_raw = f"(REPORTDATE>='{start_date}')(REPORTDATE<='{end_date}')"
    filter_escaped = quote(filter_raw, safe="")

    client = RetryClient(max_retry=args.max_retry, timeout_sec=45)

    first_page = get_financial_page_data(client, 1, filter_escaped)
    all_rows = list(first_page["Data"])
    for page_number in range(2, first_page["Pages"] + 1):
        page = get_financial_page_data(client, page_number, filter_escaped)
        all_rows.extend(page["Data"])

    normalized = normalize_rows(all_rows)
    latest_cum_rows = pick_latest_cum_rows(normalized)
    single_quarter_rows = build_single_quarter_rows(latest_cum_rows)

    current_rows = [
        row
        for row in single_quarter_rows
        if row["Year"] == args.target_year
        and row["Quarter"] == args.target_quarter
        and notice_window_start <= row["QuarterNoticeDate"] <= run_time
    ]

    prev_map = {
        row["SecurityCode"]: row
        for row in single_quarter_rows
        if row["Year"] == args.target_year - 1 and row["Quarter"] == args.target_quarter
    }

    try:
        market_map = get_market_snapshot_map(client)
    except Exception as exc:  # noqa: BLE001
        print(f"WARNING: 获取市值/上市日期/行业快照失败，将输出空值：{exc}")
        market_map = {}

    result_rows: list[dict[str, Any]] = []
    for cur in current_rows:
        prev = prev_map.get(cur["SecurityCode"])
        if prev is None:
            continue
        prev_revenue = prev.get("SingleRevenue")
        if prev_revenue in (None, 0):
            continue

        yoy = ((cur["SingleRevenue"] - prev_revenue) / prev_revenue) * 100
        if yoy <= args.growth_threshold:
            continue

        industry_l3 = ""
        market_cap = None
        listing_date_text = ""
        snap = market_map.get(cur["SecurityCode"])
        if snap is not None:
            industry_l3 = str(snap.get("IndustryL3") or "")
            market_cap = snap.get("TotalMarketCap")
            listing_date = snap.get("ListingDate")
            if listing_date is not None:
                listing_date_text = listing_date.strftime("%Y-%m-%d")

        result_rows.append(
            {
                "股票代码": cur["SecurityCode"],
                "证券代码": cur["SecuCode"],
                "股票简称": cur["Name"],
                "所属市场": cur["TradeMarket"],
                "所属三级行业": industry_l3,
                "年份": args.target_year,
                "季度": args.target_quarter,
                "报告类型": cur["QuarterReportType"],
                "季度报告期": cur["QuarterReportDate"].strftime("%Y-%m-%d"),
                "季度报发布时间": cur["QuarterNoticeDate"].strftime("%Y-%m-%d %H:%M:%S"),
                "公司总市值范围": get_market_cap_range(market_cap),
                "上市日期": listing_date_text,
                "单季度营收同比增速_百分比": round(yoy, 2),
            }
        )

    result_rows = sorted(
        result_rows,
        key=lambda row: row["单季度营收同比增速_百分比"],
        reverse=True,
    )

    summary_rows = [
        {"指标": "执行时间", "值": run_time.strftime("%Y-%m-%d %H:%M:%S")},
        {"指标": "目标年份", "值": args.target_year},
        {"指标": "目标季度", "值": args.target_quarter},
        {
            "指标": "公告发布时间窗口",
            "值": (
                f"最近{args.notice_within_hours}小时"
                f"（{notice_window_start.strftime('%Y-%m-%d %H:%M:%S')} ~ {run_time.strftime('%Y-%m-%d %H:%M:%S')}）"
            ),
        },
        {"指标": "筛选条件", "值": f"单季度营收同比增速 > {args.growth_threshold}%"},
        {"指标": "接口总记录数", "值": first_page["Count"]},
        {"指标": "A股累计营收记录数", "值": len(normalized)},
        {"指标": "24小时内目标季度公司数", "值": len(current_rows)},
        {"指标": "命中公司数", "值": len(result_rows)},
    ]

    output_dir = Path(args.output_dir) if args.output_dir else Path(__file__).resolve().parent.parent / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    stamp = run_time.strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / (
        f"a_share_{args.target_year}{args.target_quarter.lower()}_"
        f"{args.notice_within_hours}h_once_result_{stamp}.xlsx"
    )

    template_path = Path(__file__).resolve().parent.parent / "docs" / "templates" / "output_template.xlsx"
    if template_path.exists():
        try:
            workbook = build_workbook_with_template(template_path, summary_rows, result_rows)
        except Exception:  # noqa: BLE001
            workbook = build_workbook_fallback(summary_rows, result_rows)
    else:
        workbook = build_workbook_fallback(summary_rows, result_rows)

    workbook.save(output_file)

    payload = {
        "success": True,
        "run_time": run_time.strftime("%Y-%m-%d %H:%M:%S"),
        "target_quarter": f"{args.target_year}{args.target_quarter}",
        "matched_companies": len(result_rows),
        "output_excel": str(output_file.resolve()),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
