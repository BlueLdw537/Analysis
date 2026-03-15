from __future__ import annotations

import argparse
import json
import math
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

CLIST_HOSTS = [
    "https://push2.eastmoney.com",
    "https://82.push2.eastmoney.com",
    "https://88.push2.eastmoney.com",
]

QUOTE_HOSTS = [
    "https://push2.eastmoney.com",
    "https://push2his.eastmoney.com",
    "https://63.push2his.eastmoney.com",
    "https://91.push2his.eastmoney.com",
]

RESULT_HEADERS = [
    "股票代码",
    "证券代码",
    "股票简称",
    "所属市场",
    "所属三级行业",
    "年份",
    "季度",
    "报告类型",
    "季度报告期",
    "季度报发布时间",
    "公司总市值范围",
    "上市日期",
    "单季度营收同比增速_百分比",
]


class RetryClient:
    def __init__(self, max_retry: int, timeout_sec: int = 45) -> None:
        self.max_retry = max(1, int(max_retry))
        self.timeout_sec = timeout_sec
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/122.0.0.0 Safari/537.36"
                ),
                "Accept": "application/json, text/plain, */*",
            }
        )

    def get_json(self, url: str, timeout_sec: int | None = None) -> dict[str, Any]:
        timeout = timeout_sec or self.timeout_sec
        last_exc: Exception | None = None
        for attempt in range(1, self.max_retry + 1):
            try:
                response = self.session.get(url, timeout=timeout)
                response.raise_for_status()
                payload = response.json()
                if isinstance(payload, dict):
                    return payload
                raise RuntimeError(f"Unexpected JSON payload from {url}")
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if attempt >= self.max_retry:
                    break
                wait_sec = 2**attempt
                print(f"WARNING: 连接失败，第 {attempt} 次重试，等待 {wait_sec} 秒: {exc}")
                time.sleep(wait_sec)
        raise RuntimeError(f"Request failed after retries: {url}") from last_exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="A股24小时公告窗口营收同比扫描（Python v2）")
    parser.add_argument("--target-year", "--TargetYear", dest="target_year", type=int, default=2025)
    parser.add_argument("--target-quarter", "--TargetQuarter", dest="target_quarter", choices=QUARTERS, default="Q4")
    parser.add_argument("--growth-threshold", "--GrowthThreshold", dest="growth_threshold", type=float, default=20.0)
    parser.add_argument(
        "--notice-within-hours",
        "--NoticeWithinHours",
        dest="notice_within_hours",
        type=int,
        default=24,
        help="公告发布时间窗口（小时），范围 1~168",
    )
    parser.add_argument("--output-dir", "--OutputDir", dest="output_dir", default=r"D:\codex\output")
    parser.add_argument("--max-retry", "--MaxRetry", dest="max_retry", type=int, default=4)
    return parser.parse_args()


def parse_numeric(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    raw = str(value).strip()
    if not raw or raw in {"--", "-", "None", "null"}:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def get_quarter_from_month(month: int) -> str | None:
    return {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(month)


def get_single_quarter_revenue(cum_by_quarter: dict[str, float], quarter: str) -> float | None:
    if quarter == "Q1":
        return cum_by_quarter.get("Q1")
    if quarter == "Q2" and "Q2" in cum_by_quarter and "Q1" in cum_by_quarter:
        return cum_by_quarter["Q2"] - cum_by_quarter["Q1"]
    if quarter == "Q3" and "Q3" in cum_by_quarter and "Q2" in cum_by_quarter:
        return cum_by_quarter["Q3"] - cum_by_quarter["Q2"]
    if quarter == "Q4" and "Q4" in cum_by_quarter and "Q3" in cum_by_quarter:
        return cum_by_quarter["Q4"] - cum_by_quarter["Q3"]
    return None


def get_market_cap_range(market_cap: float | None) -> str:
    if market_cap is None or market_cap < 0:
        return ""
    if market_cap < 10_000_000_000:
        return "0-100亿"
    if market_cap < 50_000_000_000:
        return "100-500亿"
    if market_cap < 100_000_000_000:
        return "500-1000亿"
    return "1000亿以上"


def first_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                return item
    return {}


def get_market_prefix(secu_code: str, security_code: str) -> str:
    upper = (secu_code or "").upper()
    if upper.endswith(".SH"):
        return "SH"
    if upper.endswith(".SZ"):
        return "SZ"
    if upper.endswith(".BJ"):
        return "BJ"
    if security_code.startswith(("6", "9")):
        return "SH"
    if security_code.startswith(("0", "3")):
        return "SZ"
    return "BJ"


def get_secid(security_code: str, secu_code: str) -> str:
    upper = (secu_code or "").upper()
    if upper.endswith(".SH"):
        return f"1.{security_code}"
    if security_code.startswith(("60", "68", "90")):
        return f"1.{security_code}"
    return f"0.{security_code}"


def fetch_company_survey_profile(client: RetryClient, security_code: str, secu_code: str) -> dict[str, Any]:
    market_prefix = get_market_prefix(secu_code, security_code)
    url = f"https://emweb.securities.eastmoney.com/PC_HSF10/CompanySurvey/PageAjax?code={market_prefix}{security_code}"
    try:
        payload = client.get_json(url, timeout_sec=30)
    except Exception:  # noqa: BLE001
        return {"IndustryL3": "", "ListingDate": None}

    jbzl = first_dict(payload.get("jbzl"))
    fxxg = first_dict(payload.get("fxxg"))

    industry_l3 = ""
    chain = str(jbzl.get("EM2016") or "").strip()
    if chain:
        industry_l3 = chain.split("-")[-1].strip()

    listing_date = parse_datetime(fxxg.get("LISTING_DATE"))
    return {"IndustryL3": industry_l3, "ListingDate": listing_date}


def fetch_quote_profile(client: RetryClient, security_code: str, secu_code: str) -> dict[str, Any]:
    secid = get_secid(security_code, secu_code)
    for host in QUOTE_HOSTS:
        url = (
            f"{host}/api/qt/stock/get"
            f"?secid={secid}"
            "&fields=f57,f58,f116,f189,f127,f100"
        )
        try:
            payload = client.get_json(url, timeout_sec=30)
        except Exception:  # noqa: BLE001
            continue

        data = payload.get("data") or {}
        if not isinstance(data, dict):
            data = {}

        market_cap = parse_numeric(data.get("f116"))
        if market_cap is not None and market_cap < 0:
            market_cap = None

        industry_l3 = str(data.get("f127") or data.get("f100") or "").strip()
        if industry_l3 == "-":
            industry_l3 = ""

        listing_date = parse_datetime(data.get("f189"))

        if market_cap is None and listing_date is None and not industry_l3:
            continue
        return {
            "TotalMarketCap": market_cap,
            "IndustryL3": industry_l3,
            "ListingDate": listing_date,
        }

    return {"TotalMarketCap": None, "IndustryL3": "", "ListingDate": None}


def fetch_listing_date_from_ipo(client: RetryClient, security_code: str) -> datetime | None:
    base_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
    encoded_filter = quote(f'(SECURITY_CODE="{security_code}")', safe="")
    query = (
        "reportName=RPTA_APP_IPOAPPLY"
        "&columns=SECURITY_CODE,LISTING_DATE"
        "&source=WEB&client=WEB&pageNumber=1&pageSize=1"
        f"&filter={encoded_filter}"
    )
    url = f"{base_url}?{query}"
    try:
        payload = client.get_json(url, timeout_sec=30)
    except Exception:  # noqa: BLE001
        return None

    data = (payload.get("result") or {}).get("data") or []
    if not isinstance(data, list) or not data:
        return None
    first = data[0] if isinstance(data[0], dict) else {}
    return parse_datetime(first.get("LISTING_DATE"))


def get_supplement_profile(
    client: RetryClient,
    security_code: str,
    secu_code: str,
    cache: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    cached = cache.get(security_code)
    if cached is not None:
        return cached

    industry_l3 = ""
    market_cap: float | None = None
    listing_date: datetime | None = None

    survey = fetch_company_survey_profile(client, security_code, secu_code)
    if survey.get("IndustryL3"):
        industry_l3 = str(survey.get("IndustryL3") or "")
    if survey.get("ListingDate") is not None:
        listing_date = survey.get("ListingDate")

    quote_profile = fetch_quote_profile(client, security_code, secu_code)
    if market_cap is None and quote_profile.get("TotalMarketCap") is not None:
        market_cap = quote_profile.get("TotalMarketCap")
    if not industry_l3 and quote_profile.get("IndustryL3"):
        industry_l3 = str(quote_profile.get("IndustryL3") or "")
    if listing_date is None and quote_profile.get("ListingDate") is not None:
        listing_date = quote_profile.get("ListingDate")

    if listing_date is None:
        listing_date = fetch_listing_date_from_ipo(client, security_code)

    profile = {
        "TotalMarketCap": market_cap,
        "IndustryL3": industry_l3,
        "ListingDate": listing_date,
    }
    cache[security_code] = profile
    return profile


def get_financial_page_data(client: RetryClient, page_number: int, filter_expr: str) -> dict[str, Any]:
    base_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
    encoded_filter = quote(filter_expr, safe="")
    query = (
        "reportName=RPT_LICO_FN_CPD"
        "&columns=SECURITY_CODE,SECUCODE,SECURITY_NAME_ABBR,SECURITY_TYPE,TRADE_MARKET,"
        "DATATYPE,REPORTDATE,NOTICE_DATE,TOTAL_OPERATE_INCOME"
        "&source=WEB"
        "&client=WEB"
        f"&pageNumber={page_number}"
        "&pageSize=500"
        f"&filter={encoded_filter}"
    )
    response = client.get_json(f"{base_url}?{query}")
    result = response.get("result") or {}
    data = result.get("data") or []
    if not isinstance(data, list):
        data = []
    return {
        "Data": data,
        "Pages": int(result.get("pages") or 0),
        "Count": int(result.get("count") or 0),
    }


def get_market_snapshot_map(client: RetryClient) -> dict[str, dict[str, Any]]:
    errors: list[str] = []

    for host in CLIST_HOSTS:
        snapshot: dict[str, dict[str, Any]] = {}
        try:
            base_url = f"{host}/api/qt/clist/get"
            page_size = 500
            query = (
                f"pz={page_size}&po=1&np=1"
                "&ut=bd1d9ddb04089700cf9c27f6f7426281"
                "&fltt=2&invt=2&fid=f3"
                "&fs=m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23"
                "&fields=f12,f20,f26,f100"
            )

            def add_diff(diff: list[dict[str, Any]] | None) -> None:
                if not diff:
                    return
                for row in diff:
                    code = str(row.get("f12") or "").strip()
                    if not code:
                        continue

                    market_cap = parse_numeric(row.get("f20"))
                    if market_cap is not None and market_cap < 0:
                        market_cap = None

                    listing_date = None
                    listing_raw = str(row.get("f26") or "").strip()
                    if len(listing_raw) == 8 and listing_raw.isdigit():
                        try:
                            listing_date = datetime.strptime(listing_raw, "%Y%m%d")
                        except ValueError:
                            listing_date = None

                    industry_l3 = ""
                    industry_text = str(row.get("f100") or "").strip()
                    if industry_text and industry_text != "-":
                        industry_l3 = industry_text

                    snapshot[code] = {
                        "TotalMarketCap": market_cap,
                        "ListingDate": listing_date,
                        "IndustryL3": industry_l3,
                    }

            first_resp = client.get_json(f"{base_url}?pn=1&{query}")
            first_data = first_resp.get("data") or {}
            total = int(first_data.get("total") or 0)
            pages = int(math.ceil(total / page_size)) if total > 0 else 0
            pages = max(pages, 1)

            add_diff(first_data.get("diff"))
            for page_number in range(2, pages + 1):
                resp = client.get_json(f"{base_url}?pn={page_number}&{query}")
                add_diff((resp.get("data") or {}).get("diff"))

            if snapshot:
                print(f"INFO: 市场快照来源 {host}")
                return snapshot
            raise RuntimeError("empty snapshot payload")
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{host}: {exc}")

    print(
        "WARNING: 获取市场快照失败，所属三级行业/市值/上市日期将为空："
        + " ; ".join(errors)
    )
    return {}


def normalize_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in raw_rows:
        if row.get("SECURITY_TYPE") != "A股":
            continue

        income = parse_numeric(row.get("TOTAL_OPERATE_INCOME"))
        if income is None:
            continue

        report_date = parse_datetime(row.get("REPORTDATE"))
        notice_date = parse_datetime(row.get("NOTICE_DATE"))
        if report_date is None or notice_date is None:
            continue

        quarter = get_quarter_from_month(report_date.month)
        if quarter is None:
            continue

        normalized.append(
            {
                "SecurityCode": str(row.get("SECURITY_CODE") or ""),
                "SecuCode": str(row.get("SECUCODE") or ""),
                "Name": str(row.get("SECURITY_NAME_ABBR") or ""),
                "TradeMarket": str(row.get("TRADE_MARKET") or ""),
                "ReportType": str(row.get("DATATYPE") or ""),
                "ReportDate": report_date,
                "NoticeDate": notice_date,
                "Year": int(report_date.year),
                "Quarter": quarter,
                "CumRevenue": float(income),
            }
        )
    return normalized


def pick_latest_cum_rows(normalized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[tuple[str, int, str], dict[str, Any]] = {}
    for row in normalized_rows:
        key = (row["SecurityCode"], row["Year"], row["Quarter"])
        old = latest.get(key)
        if old is None or row["NoticeDate"] > old["NoticeDate"]:
            latest[key] = row
    return list(latest.values())


def build_single_quarter_rows(latest_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in latest_rows:
        grouped[(row["SecurityCode"], row["Year"])].append(row)

    output: list[dict[str, Any]] = []
    for rows in grouped.values():
        cum_by_quarter: dict[str, float] = {}
        row_by_quarter: dict[str, dict[str, Any]] = {}
        for item in rows:
            cum_by_quarter[item["Quarter"]] = item["CumRevenue"]
            row_by_quarter[item["Quarter"]] = item

        for quarter in QUARTERS:
            single = get_single_quarter_revenue(cum_by_quarter, quarter)
            if single is None or quarter not in row_by_quarter:
                continue
            q_row = row_by_quarter[quarter]
            output.append(
                {
                    "SecurityCode": q_row["SecurityCode"],
                    "SecuCode": q_row["SecuCode"],
                    "Name": q_row["Name"],
                    "TradeMarket": q_row["TradeMarket"],
                    "Year": int(q_row["Year"]),
                    "Quarter": quarter,
                    "QuarterReportType": q_row["ReportType"],
                    "QuarterReportDate": q_row["ReportDate"],
                    "QuarterNoticeDate": q_row["NoticeDate"],
                    "SingleRevenue": float(single),
                }
            )
    return output


def auto_fit_worksheet(worksheet: Any) -> None:
    for idx, col in enumerate(worksheet.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)


def write_values(worksheet: Any, rows: list[dict[str, Any]], headers: list[str]) -> None:
    max_cols = len(headers)
    target_last_row = max(2, len(rows) + 1)
    clear_to_row = max(worksheet.max_row, target_last_row)

    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)

    for row_idx in range(2, clear_to_row + 1):
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    if not rows:
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=2, column=col_idx, value="")

    auto_fit_worksheet(worksheet)


def upsert_table(
    worksheet: Any,
    table_name: str,
    max_cols: int,
    data_rows: int,
    style_name: str = "TableStyleMedium6",
) -> None:
    table_last_row = max(2, data_rows + 1)
    table_ref = f"A1:{get_column_letter(max_cols)}{table_last_row}"
    if table_name in worksheet.tables:
        table = worksheet.tables[table_name]
        table.ref = table_ref
        return

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def build_workbook(summary_rows: list[dict[str, Any]], result_rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    ws_result = workbook.create_sheet("Result")

    summary_headers = ["指标", "值"]
    write_values(ws_summary, summary_rows, summary_headers)
    write_values(ws_result, result_rows, RESULT_HEADERS)
    upsert_table(ws_summary, "SummaryTable", len(summary_headers), len(summary_rows))
    upsert_table(ws_result, "ResultTable", len(RESULT_HEADERS), len(result_rows))
    return workbook


def main() -> None:
    args = parse_args()
    if not (1 <= args.notice_within_hours <= 168):
        raise ValueError("--notice-within-hours 必须在 1~168 范围内。")

    run_time = datetime.now()
    notice_window_start = run_time - timedelta(hours=args.notice_within_hours)

    start_date = f"{args.target_year - 1}-01-01"
    end_date = f"{args.target_year}-12-31"
    filter_raw = f"(REPORTDATE>='{start_date}')(REPORTDATE<='{end_date}')"

    client = RetryClient(max_retry=args.max_retry, timeout_sec=45)
    first_page = get_financial_page_data(client, page_number=1, filter_expr=filter_raw)

    all_rows: list[dict[str, Any]] = list(first_page["Data"])
    for page_number in range(2, first_page["Pages"] + 1):
        page_data = get_financial_page_data(client, page_number=page_number, filter_expr=filter_raw)
        all_rows.extend(page_data["Data"])

    normalized = normalize_rows(all_rows)
    latest_cum_rows = pick_latest_cum_rows(normalized)
    single_quarter_rows = build_single_quarter_rows(latest_cum_rows)

    current_rows = [
        row
        for row in single_quarter_rows
        if row["Year"] == args.target_year
        and row["Quarter"] == args.target_quarter
        and notice_window_start <= row["QuarterNoticeDate"] <= run_time
    ]
    prev_map = {
        row["SecurityCode"]: row
        for row in single_quarter_rows
        if row["Year"] == (args.target_year - 1) and row["Quarter"] == args.target_quarter
    }

    market_map = get_market_snapshot_map(client)
    supplement_cache: dict[str, dict[str, Any]] = {}
    supplement_hit_count = 0

    result_rows: list[dict[str, Any]] = []
    for cur in current_rows:
        prev = prev_map.get(cur["SecurityCode"])
        if prev is None:
            continue
        prev_revenue = prev.get("SingleRevenue")
        if prev_revenue in (None, 0):
            continue

        yoy = ((cur["SingleRevenue"] - prev_revenue) / prev_revenue) * 100.0
        if yoy <= args.growth_threshold:
            continue

        industry_l3 = ""
        market_cap: float | None = None
        listing_date: datetime | None = None
        snap = market_map.get(cur["SecurityCode"])
        if snap:
            industry_l3 = str(snap.get("IndustryL3") or "")
            market_cap = parse_numeric(snap.get("TotalMarketCap"))
            listing_date = parse_datetime(snap.get("ListingDate"))

        supplement_used = False
        if market_cap is None or not industry_l3 or listing_date is None:
            supplement = get_supplement_profile(
                client=client,
                security_code=cur["SecurityCode"],
                secu_code=cur["SecuCode"],
                cache=supplement_cache,
            )
            if market_cap is None and supplement.get("TotalMarketCap") is not None:
                market_cap = supplement.get("TotalMarketCap")
                supplement_used = True
            if not industry_l3 and supplement.get("IndustryL3"):
                industry_l3 = str(supplement.get("IndustryL3") or "")
                supplement_used = True
            if listing_date is None and supplement.get("ListingDate") is not None:
                listing_date = supplement.get("ListingDate")
                supplement_used = True

        if supplement_used:
            supplement_hit_count += 1

        listing_date_text = ""
        if listing_date is not None:
            listing_date_text = listing_date.strftime("%Y-%m-%d")

        result_rows.append(
            {
                "股票代码": cur["SecurityCode"],
                "证券代码": cur["SecuCode"],
                "股票简称": cur["Name"],
                "所属市场": cur["TradeMarket"],
                "所属三级行业": industry_l3,
                "年份": args.target_year,
                "季度": args.target_quarter,
                "报告类型": cur["QuarterReportType"],
                "季度报告期": cur["QuarterReportDate"].strftime("%Y-%m-%d"),
                "季度报发布时间": cur["QuarterNoticeDate"].strftime("%Y-%m-%d %H:%M:%S"),
                "公司总市值范围": get_market_cap_range(market_cap),
                "上市日期": listing_date_text,
                "单季度营收同比增速_百分比": round(yoy, 2),
            }
        )

    result_rows.sort(key=lambda row: row["单季度营收同比增速_百分比"], reverse=True)

    summary_rows = [
        {"指标": "执行时间", "值": run_time.strftime("%Y-%m-%d %H:%M:%S")},
        {"指标": "目标年份", "值": args.target_year},
        {"指标": "目标季度", "值": args.target_quarter},
        {
            "指标": "公告发布时间窗口",
            "值": (
                f"最近{args.notice_within_hours}小时"
                f"（{notice_window_start.strftime('%Y-%m-%d %H:%M:%S')} ~ {run_time.strftime('%Y-%m-%d %H:%M:%S')}）"
            ),
        },
        {"指标": "筛选条件", "值": f"单季度营收同比增速 > {args.growth_threshold}%"},
        {"指标": "接口总记录数", "值": first_page["Count"]},
        {"指标": "A股记录数", "值": len(normalized)},
        {"指标": "24小时目标季度公司数", "值": len(current_rows)},
        {"指标": "补充公开数据命中数", "值": supplement_hit_count},
        {"指标": "命中公司数", "值": len(result_rows)},
    ]

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = run_time.strftime("%Y%m%d_%H%M%S")
    quarter_token = f"{args.target_year}{args.target_quarter.lower()}"
    output_path = output_dir / f"a_share_{quarter_token}_24h_once_result_{stamp}.xlsx"

    workbook = build_workbook(summary_rows, result_rows)
    workbook.save(output_path)

    payload = {
        "success": True,
        "run_time": run_time.strftime("%Y-%m-%d %H:%M:%S"),
        "target_quarter": f"{args.target_year}{args.target_quarter}",
        "matched_companies": len(result_rows),
        "output_excel": str(output_path.resolve()),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
