
from __future__ import annotations

import argparse
import json
import math
import re
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


RESULT_HEADERS = [
    "股票代码",
    "证券代码",
    "股票简称",
    "所属市场",
    "所属板块",
    "所属三级行业",
    "上市日期",
    "上市年限_年",
    "是否ST",
    "区间起始交易日",
    "区间结束交易日",
    "区间起始收盘价",
    "区间结束收盘价",
    "区间涨跌幅_百分比",
    "区间振幅_百分比",
    "区间交易天数",
    "区间最高价",
    "区间最低价",
    "最新价",
    "最新涨跌幅_百分比",
    "换手率_百分比",
    "成交额_元",
    "总市值_元",
    "流通市值_元",
    "市值区间",
    "PE_TTM",
    "PB",
    "60日涨跌幅_百分比",
    "年初至今涨跌幅_百分比",
    "数据源",
]

CLIST_HOSTS = [
    "https://push2.eastmoney.com",
    "https://82.push2.eastmoney.com",
    "https://88.push2.eastmoney.com",
]

HISTORY_HOSTS = [
    "https://push2his.eastmoney.com",
    "https://63.push2his.eastmoney.com",
    "https://91.push2his.eastmoney.com",
]


class RetryClient:
    def __init__(self, max_retry: int, timeout_sec: int = 30, request_interval_ms: int = 0) -> None:
        self.max_retry = max_retry
        self.timeout_sec = timeout_sec
        self.request_interval_ms = max(0, request_interval_ms)
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            }
        )
        self._last_request_ts = 0.0

    def _wait_interval(self) -> None:
        if self.request_interval_ms <= 0:
            return
        gap = (time.time() - self._last_request_ts) * 1000
        if gap < self.request_interval_ms:
            time.sleep((self.request_interval_ms - gap) / 1000)

    def _mark_request(self) -> None:
        self._last_request_ts = time.time()

    def get_json(self, url: str, timeout_sec: int | None = None) -> dict[str, Any]:
        timeout = timeout_sec or self.timeout_sec
        last_exc: Exception | None = None

        for attempt in range(1, self.max_retry + 1):
            try:
                self._wait_interval()
                response = self.session.get(url, timeout=timeout)
                self._mark_request()
                response.raise_for_status()
                return response.json()
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if attempt >= self.max_retry:
                    break
                time.sleep(2 ** attempt)

        raise RuntimeError(f"Request failed after retries: {url}") from last_exc


def parse_args() -> argparse.Namespace:
    now_year = datetime.now().year

    parser = argparse.ArgumentParser(description="A股区间涨跌幅筛选（支持主/备行情接口自动切换）")
    parser.add_argument("--start-year", "--StartYear", dest="start_year", type=int, default=now_year - 1, help="起始年份")
    parser.add_argument("--end-year", "--EndYear", dest="end_year", type=int, default=now_year, help="结束年份")
    parser.add_argument(
        "--change-threshold-pct",
        "--ChangeThresholdPct",
        dest="change_threshold_pct",
        type=float,
        default=100.0,
        help="涨跌幅阈值（百分比，默认100）",
    )
    parser.add_argument(
        "--direction",
        "--Direction",
        dest="direction",
        choices=["rise", "fall", "abs"],
        default="rise",
        help="方向：rise/fall/abs",
    )
    parser.add_argument("--top-n", "--TopN", dest="top_n", type=int, default=0, help="限制输出数量，0表示不限")
    parser.add_argument(
        "--max-stocks",
        "--MaxStocks",
        dest="max_stocks",
        type=int,
        default=0,
        help="处理股票数量上限，0表示不限",
    )
    parser.add_argument(
        "--request-interval-ms",
        "--RequestIntervalMs",
        dest="request_interval_ms",
        type=int,
        default=120,
        help="请求间隔毫秒，默认120",
    )
    parser.add_argument("--output-path", "--OutputPath", dest="output_path", default="", help="输出Excel完整路径")
    parser.add_argument("--max-retry", "--MaxRetry", dest="max_retry", type=int, default=4, help="接口重试次数")
    parser.add_argument("--timeout-sec", "--TimeoutSec", dest="timeout_sec", type=int, default=30, help="请求超时秒数")
    return parser.parse_args()


def parse_float(value: Any) -> float | None:
    try:
        if value in (None, "", "-", "--", "null", "None"):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_int(value: Any) -> int | None:
    try:
        if value in (None, "", "-", "--", "null", "None"):
            return None
        return int(str(value))
    except (TypeError, ValueError):
        return None


def parse_yyyymmdd(value: Any) -> datetime | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw in {"-", "--", "null", "None"}:
        return None

    if re.fullmatch(r"\d{8}", raw):
        try:
            return datetime.strptime(raw, "%Y%m%d")
        except ValueError:
            return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue

    return None


def scaled_pct(value: Any) -> float | None:
    num = parse_float(value)
    if num is None:
        return None
    return num / 100.0


def normalize_threshold(value: float) -> float:
    return abs(float(value))


def is_st_stock(name: str) -> bool:
    upper = (name or "").upper().replace(" ", "")
    return "ST" in upper


def is_a_share_code(code: str) -> bool:
    if not re.fullmatch(r"\d{6}", code):
        return False
    if code.startswith(("200", "900")):
        return False
    return code.startswith(
        (
            "000",
            "001",
            "002",
            "003",
            "300",
            "301",
            "600",
            "601",
            "603",
            "605",
            "688",
            "689",
            "4",
            "8",
            "9",
        )
    )


def market_from_code_or_text(code: str, trade_market: str, secucode: str) -> str:
    if trade_market:
        if "沪" in trade_market:
            return "沪A"
        if "深" in trade_market:
            return "深A"
        if "北" in trade_market:
            return "北A"

    upper = secucode.upper()
    if upper.endswith(".SH"):
        return "沪A"
    if upper.endswith(".SZ"):
        return "深A"
    if upper.endswith(".BJ"):
        return "北A"

    if code.startswith(("600", "601", "603", "605", "688", "689")):
        return "沪A"
    if code.startswith(("4", "8", "9")):
        return "北A"
    return "深A"


def board_from_code_or_text(code: str, board_name: str) -> str:
    if board_name and board_name not in {"-", "--"}:
        return board_name
    if code.startswith(("688", "689")):
        return "科创板"
    if code.startswith(("300", "301")):
        return "创业板"
    if code.startswith(("4", "8", "9")):
        return "北交所"
    if code.startswith(("600", "601", "603", "605")):
        return "沪市主板"
    if code.startswith(("000", "001", "002", "003")):
        return "深市主板"
    return ""

def get_market_cap_bucket(total_market_cap: float | None) -> str:
    if total_market_cap is None or total_market_cap <= 0:
        return ""
    if total_market_cap < 10_000_000_000:
        return "0-100亿"
    if total_market_cap < 50_000_000_000:
        return "100-500亿"
    if total_market_cap < 100_000_000_000:
        return "500-1000亿"
    if total_market_cap < 300_000_000_000:
        return "1000-3000亿"
    return "3000亿以上"


def build_secid(code: str, secucode: str = "") -> str:
    upper = secucode.upper()
    if upper.endswith(".SH"):
        return f"1.{code}"
    if upper.endswith(".SZ") or upper.endswith(".BJ"):
        return f"0.{code}"
    if code.startswith(("600", "601", "603", "605", "688", "689", "900")):
        return f"1.{code}"
    return f"0.{code}"


def format_threshold_for_filename(value: float) -> str:
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return text.replace(".", "p")


def fetch_universe_from_clist(client: RetryClient, max_stocks: int) -> tuple[list[dict[str, Any]], str]:
    fields = ",".join(
        [
            "f12",
            "f14",
            "f13",
            "f100",
            "f127",
            "f128",
            "f26",
            "f2",
            "f3",
            "f8",
            "f6",
            "f20",
            "f21",
            "f9",
            "f23",
            "f24",
            "f25",
            "f119",
            "f120",
            "f162",
            "f167",
        ]
    )

    query = (
        "po=1&np=1&ut=bd1d9ddb04089700cf9c27f6f7426281&fltt=2&invt=2&fid=f12"
        "&fs=m:0+t:6,m:0+t:80,m:0+t:81+s:2048,m:1+t:2,m:1+t:23"
        "&pz=500"
        f"&fields={fields}"
    )

    errors: list[str] = []
    for host in CLIST_HOSTS:
        try:
            first_url = f"{host}/api/qt/clist/get?pn=1&{query}"
            first = client.get_json(first_url)
            data = first.get("data") or {}
            total = int(data.get("total") or 0)
            pages = int(math.ceil(total / 500)) if total > 0 else 0
            pages = max(1, pages)

            results: list[dict[str, Any]] = []

            def ingest(diff: list[dict[str, Any]] | None) -> None:
                if not diff:
                    return
                for row in diff:
                    code = str(row.get("f12") or "").strip()
                    if not is_a_share_code(code):
                        continue

                    name = str(row.get("f14") or "").strip()
                    market_id = parse_int(row.get("f13"))
                    secucode = f"{code}.{'SH' if market_id == 1 else 'SZ'}"
                    if code.startswith(("4", "8", "9")):
                        secucode = f"{code}.BJ"

                    trade_market = ""
                    if market_id == 1:
                        trade_market = "沪A"
                    elif market_id == 0:
                        trade_market = "深A"
                    if code.startswith(("4", "8", "9")):
                        trade_market = "北A"

                    listing_date = parse_yyyymmdd(row.get("f26"))
                    pe_ttm = parse_float(row.get("f9"))
                    if pe_ttm is None:
                        pe_ttm = scaled_pct(row.get("f162"))

                    pb = parse_float(row.get("f23"))
                    if pb is None:
                        pb = scaled_pct(row.get("f167"))

                    pct_60d = parse_float(row.get("f24"))
                    if pct_60d is None:
                        pct_60d = scaled_pct(row.get("f119"))

                    pct_ytd = parse_float(row.get("f25"))
                    if pct_ytd is None:
                        pct_ytd = scaled_pct(row.get("f120"))

                    results.append(
                        {
                            "code": code,
                            "secucode": secucode,
                            "name": name,
                            "trade_market": trade_market,
                            "board": board_from_code_or_text(code, str(row.get("f128") or "")),
                            "industry_l3": str(row.get("f127") or row.get("f100") or "").strip(),
                            "listing_date": listing_date,
                            "latest_price": parse_float(row.get("f2")),
                            "latest_change_pct": parse_float(row.get("f3")),
                            "turnover_pct": parse_float(row.get("f8")),
                            "amount": parse_float(row.get("f6")),
                            "total_market_cap": parse_float(row.get("f20")),
                            "float_market_cap": parse_float(row.get("f21")),
                            "pe_ttm": pe_ttm,
                            "pb": pb,
                            "pct_60d": pct_60d,
                            "pct_ytd": pct_ytd,
                            "profile_source": f"clist@{host}",
                        }
                    )

                    if max_stocks > 0 and len(results) >= max_stocks:
                        return

            ingest(data.get("diff"))
            if max_stocks == 0 or len(results) < max_stocks:
                for page_number in range(2, pages + 1):
                    resp = client.get_json(f"{host}/api/qt/clist/get?pn={page_number}&{query}")
                    ingest((resp.get("data") or {}).get("diff"))
                    if max_stocks > 0 and len(results) >= max_stocks:
                        break

            if not results:
                raise RuntimeError("clist返回空股票池")
            return results, f"clist@{host}"
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{host}: {exc}")

    raise RuntimeError(" ; ".join(errors))


def fetch_universe_from_datacenter(client: RetryClient, end_year: int, max_stocks: int) -> tuple[list[dict[str, Any]], str]:
    window_start_year = max(2008, end_year - 2)
    encoded_filter = quote(f"(REPORTDATE>='{window_start_year}-01-01')", safe="")

    base_url = "https://datacenter-web.eastmoney.com/api/data/v1/get"
    columns = (
        "SECURITY_CODE,SECUCODE,SECURITY_NAME_ABBR,SECURITY_TYPE,"
        "TRADE_MARKET,BOARD_NAME,REPORTDATE,NOTICE_DATE"
    )

    def fetch_page(page_number: int) -> dict[str, Any]:
        query = (
            "reportName=RPT_LICO_FN_CPD"
            f"&columns={columns}"
            "&source=WEB&client=WEB&pageSize=500"
            f"&filter={encoded_filter}&pageNumber={page_number}"
        )
        return client.get_json(f"{base_url}?{query}")

    first = fetch_page(1)
    first_result = first.get("result") or {}
    total_pages = int(first_result.get("pages") or 0)
    total_pages = max(total_pages, 1)

    latest_map: dict[str, dict[str, Any]] = {}

    def ingest(data_rows: list[dict[str, Any]] | None) -> None:
        if not data_rows:
            return
        for row in data_rows:
            if str(row.get("SECURITY_TYPE") or "") != "A股":
                continue
            code = str(row.get("SECURITY_CODE") or "").strip()
            if not is_a_share_code(code):
                continue

            notice_dt = parse_yyyymmdd(row.get("NOTICE_DATE"))
            report_dt = parse_yyyymmdd(row.get("REPORTDATE"))
            sort_dt = notice_dt or report_dt or datetime.min

            old = latest_map.get(code)
            if old is not None and sort_dt <= old["_sort_dt"]:
                continue

            latest_map[code] = {
                "code": code,
                "secucode": str(row.get("SECUCODE") or "").strip(),
                "name": str(row.get("SECURITY_NAME_ABBR") or "").strip(),
                "trade_market": str(row.get("TRADE_MARKET") or "").strip(),
                "board": str(row.get("BOARD_NAME") or "").strip(),
                "industry_l3": "",
                "listing_date": None,
                "latest_price": None,
                "latest_change_pct": None,
                "turnover_pct": None,
                "amount": None,
                "total_market_cap": None,
                "float_market_cap": None,
                "pe_ttm": None,
                "pb": None,
                "pct_60d": None,
                "pct_ytd": None,
                "profile_source": "datacenter-web",
                "_sort_dt": sort_dt,
            }

    ingest(first_result.get("data"))
    for page_number in range(2, total_pages + 1):
        resp = fetch_page(page_number)
        ingest((resp.get("result") or {}).get("data"))

    if not latest_map:
        raise RuntimeError("datacenter-web返回空股票池")

    stocks = sorted(latest_map.values(), key=lambda item: item["code"])
    for item in stocks:
        item.pop("_sort_dt", None)

    if max_stocks > 0:
        stocks = stocks[:max_stocks]
    return stocks, "datacenter-web"


def fetch_history_bars(client: RetryClient, secid: str, beg_ymd: str, end_ymd: str) -> tuple[list[dict[str, Any]], str]:
    errors: list[str] = []

    for host in HISTORY_HOSTS:
        try:
            url = (
                f"{host}/api/qt/stock/kline/get"
                f"?secid={secid}"
                "&ut=7eea3edcaed734bea9cbfc24409ed989"
                "&fields1=f1,f2,f3,f4,f5,f6"
                "&fields2=f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61"
                "&klt=101&fqt=1&lmt=10000"
                f"&beg={beg_ymd}&end={end_ymd}"
            )
            resp = client.get_json(url)
            klines = ((resp.get("data") or {}).get("klines")) or []
            bars: list[dict[str, Any]] = []

            for line in klines:
                parts = str(line).split(",")
                if len(parts) < 5:
                    continue
                date_val = parse_yyyymmdd(parts[0])
                close_val = parse_float(parts[2])
                high_val = parse_float(parts[3])
                low_val = parse_float(parts[4])
                amount_val = parse_float(parts[6]) if len(parts) > 6 else None
                turnover_val = parse_float(parts[10]) if len(parts) > 10 else None

                if date_val is None or close_val is None:
                    continue

                bars.append(
                    {
                        "date": date_val,
                        "close": close_val,
                        "high": high_val,
                        "low": low_val,
                        "amount": amount_val,
                        "turnover_pct": turnover_val,
                    }
                )

            if not bars:
                raise RuntimeError("kline为空")

            bars.sort(key=lambda item: item["date"])
            return bars, f"kline@{host}"
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{host}: {exc}")

    raise RuntimeError(" ; ".join(errors))


def fetch_stock_detail(client: RetryClient, secid: str) -> dict[str, Any]:
    fields = ",".join(
        [
            "f57",
            "f58",
            "f43",
            "f170",
            "f168",
            "f48",
            "f116",
            "f117",
            "f119",
            "f120",
            "f162",
            "f167",
            "f127",
            "f128",
            "f189",
        ]
    )
    url = (
        "https://push2.eastmoney.com/api/qt/stock/get"
        f"?secid={secid}&ut=fa5fd1943c7b386f172d6893dbfba10b&fields={fields}"
    )
    resp = client.get_json(url)
    data = resp.get("data") or {}

    return {
        "name": str(data.get("f58") or "").strip(),
        "latest_price": scaled_pct(data.get("f43")),
        "latest_change_pct": scaled_pct(data.get("f170")),
        "turnover_pct": scaled_pct(data.get("f168")),
        "amount": parse_float(data.get("f48")),
        "total_market_cap": parse_float(data.get("f116")),
        "float_market_cap": parse_float(data.get("f117")),
        "pct_60d": scaled_pct(data.get("f119")),
        "pct_ytd": scaled_pct(data.get("f120")),
        "pe_ttm": scaled_pct(data.get("f162")),
        "pb": scaled_pct(data.get("f167")),
        "industry_l3": str(data.get("f127") or "").strip(),
        "board": str(data.get("f128") or "").strip(),
        "listing_date": parse_yyyymmdd(data.get("f189")),
        "detail_source": "stock/get@push2.eastmoney.com",
    }


def need_detail_enrichment(stock: dict[str, Any]) -> bool:
    must_fields = [
        "latest_price",
        "latest_change_pct",
        "turnover_pct",
        "amount",
        "total_market_cap",
        "float_market_cap",
        "pe_ttm",
        "pb",
        "pct_60d",
        "pct_ytd",
    ]
    if any(stock.get(field) is None for field in must_fields):
        return True
    if not stock.get("industry_l3") or not stock.get("board") or stock.get("listing_date") is None:
        return True
    return False


def apply_detail(stock: dict[str, Any], detail: dict[str, Any]) -> None:
    for key, value in detail.items():
        if key == "detail_source":
            continue
        if value in (None, "", "--"):
            continue
        stock[key] = value


def compute_listing_years(listing_date: datetime | None, ref_time: datetime) -> float | None:
    if listing_date is None:
        return None
    diff_days = (ref_time.date() - listing_date.date()).days
    if diff_days < 0:
        return None
    return round(diff_days / 365.25, 2)


def matches_direction(change_pct: float, direction: str, threshold_pct: float) -> bool:
    if direction == "rise":
        return change_pct >= threshold_pct
    if direction == "fall":
        return change_pct <= -threshold_pct
    return abs(change_pct) >= threshold_pct


def sort_candidates(candidates: list[dict[str, Any]], direction: str) -> list[dict[str, Any]]:
    if direction == "rise":
        return sorted(candidates, key=lambda row: row["interval_change_pct"], reverse=True)
    if direction == "fall":
        return sorted(candidates, key=lambda row: row["interval_change_pct"])
    return sorted(candidates, key=lambda row: abs(row["interval_change_pct"]), reverse=True)


def auto_fit_worksheet(worksheet: Any) -> None:
    for idx, col in enumerate(worksheet.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)


def write_values_keep_style(worksheet: Any, rows: list[dict[str, Any]], headers: list[str]) -> None:
    max_cols = len(headers)
    target_last_row = max(2, len(rows) + 1)
    clear_to_row = max(worksheet.max_row, target_last_row)

    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)

    for row_idx in range(2, clear_to_row + 1):
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    if not rows:
        for col_idx in range(1, max_cols + 1):
            worksheet.cell(row=2, column=col_idx, value="")

    auto_fit_worksheet(worksheet)

def upsert_table(
    worksheet: Any,
    table_name: str,
    max_cols: int,
    data_rows: int,
    style_name: str = "TableStyleMedium6",
) -> None:
    table_last_row = max(2, data_rows + 1)
    table_ref = f"A1:{get_column_letter(max_cols)}{table_last_row}"

    if table_name in worksheet.tables:
        table = worksheet.tables[table_name]
        table.ref = table_ref
        return

    if worksheet.tables:
        existing = next(iter(worksheet.tables.values()))
        existing.ref = table_ref
        return

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def build_workbook_with_template(
    template_path: Path,
    summary_rows: list[dict[str, Any]],
    result_rows: list[dict[str, Any]],
) -> Workbook:
    workbook = load_workbook(template_path)
    if "Summary" not in workbook.sheetnames or "Result" not in workbook.sheetnames:
        raise ValueError("样式模板必须包含 Summary 和 Result 两个工作表。")

    summary_sheet = workbook["Summary"]
    result_sheet = workbook["Result"]

    summary_headers = ["指标", "值"]
    write_values_keep_style(summary_sheet, summary_rows, summary_headers)
    write_values_keep_style(result_sheet, result_rows, RESULT_HEADERS)

    upsert_table(summary_sheet, "SummaryTable", len(summary_headers), len(summary_rows))
    upsert_table(result_sheet, "ResultTable", len(RESULT_HEADERS), len(result_rows))

    return workbook


def build_workbook_fallback(summary_rows: list[dict[str, Any]], result_rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_headers = ["指标", "值"]
    write_values_keep_style(summary_sheet, summary_rows, summary_headers)

    result_sheet = workbook.create_sheet("Result")
    write_values_keep_style(result_sheet, result_rows, RESULT_HEADERS)

    upsert_table(summary_sheet, "SummaryTable", len(summary_headers), len(summary_rows))
    upsert_table(result_sheet, "ResultTable", len(RESULT_HEADERS), len(result_rows))

    return workbook


def to_date_text(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d")


def round_or_none(value: float | None, digits: int = 2) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def main() -> None:
    args = parse_args()

    start_year = int(args.start_year)
    end_year = int(args.end_year)
    threshold_pct = normalize_threshold(args.change_threshold_pct)
    direction = str(args.direction)
    top_n = max(0, int(args.top_n or 0))
    max_stocks = max(0, int(args.max_stocks or 0))

    current_year = datetime.now().year
    if start_year < 1990 or end_year > current_year + 1 or start_year > end_year:
        raise ValueError(f"年份参数无效，请确保 1990 <= StartYear <= EndYear <= {current_year + 1}。")

    run_time = datetime.now()
    interval_start = datetime(start_year, 1, 1)
    interval_end = datetime(end_year, 12, 31)
    beg_ymd = interval_start.strftime("%Y%m%d")
    end_ymd = interval_end.strftime("%Y%m%d")

    if args.output_path:
        output_path = Path(args.output_path)
    else:
        stamp = run_time.strftime("%Y%m%d_%H%M%S")
        threshold_tag = format_threshold_for_filename(threshold_pct)
        output_path = (
            Path(__file__).resolve().parents[2]
            / "output"
            / f"a_share_interval_{start_year}_{end_year}_{direction}_ge{threshold_tag}_{stamp}.xlsx"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    client = RetryClient(
        max_retry=max(1, int(args.max_retry)),
        timeout_sec=max(5, int(args.timeout_sec)),
        request_interval_ms=max(0, int(args.request_interval_ms)),
    )

    primary_quote_ok = True
    quote_source_note = ""
    quote_fallback_reason = ""

    try:
        stocks, quote_source_note = fetch_universe_from_clist(client, max_stocks=max_stocks)
    except Exception as exc:  # noqa: BLE001
        primary_quote_ok = False
        quote_fallback_reason = str(exc)
        stocks, quote_source_note = fetch_universe_from_datacenter(client, end_year=end_year, max_stocks=max_stocks)

    if max_stocks > 0 and len(stocks) > max_stocks:
        stocks = stocks[:max_stocks]

    universe_count = len(stocks)
    processed_count = 0
    valid_interval_count = 0
    history_error_count = 0
    insufficient_kline_count = 0

    history_source_counter: Counter[str] = Counter()
    candidates: list[dict[str, Any]] = []

    for stock in stocks:
        code = str(stock.get("code") or "").strip()
        if not code:
            continue

        processed_count += 1
        secucode = str(stock.get("secucode") or "").strip()
        secid = build_secid(code, secucode)

        try:
            bars, history_source = fetch_history_bars(client, secid=secid, beg_ymd=beg_ymd, end_ymd=end_ymd)
            history_source_counter[history_source] += 1
        except Exception:
            history_error_count += 1
            continue

        if len(bars) < 2:
            insufficient_kline_count += 1
            continue

        start_bar = bars[0]
        end_bar = bars[-1]
        start_close = parse_float(start_bar.get("close"))
        end_close = parse_float(end_bar.get("close"))
        if start_close is None or end_close is None or start_close <= 0:
            insufficient_kline_count += 1
            continue

        valid_interval_count += 1

        interval_change_pct = ((end_close - start_close) / start_close) * 100.0
        if not matches_direction(interval_change_pct, direction=direction, threshold_pct=threshold_pct):
            continue

        highs = [parse_float(bar.get("high")) for bar in bars if parse_float(bar.get("high")) is not None]
        lows = [parse_float(bar.get("low")) for bar in bars if parse_float(bar.get("low")) is not None]

        interval_high = max(highs) if highs else None
        interval_low = min(lows) if lows else None
        interval_amplitude_pct: float | None = None
        if interval_high is not None and interval_low is not None and start_close > 0:
            interval_amplitude_pct = ((interval_high - interval_low) / start_close) * 100.0

        market = market_from_code_or_text(code, str(stock.get("trade_market") or ""), secucode)
        board = board_from_code_or_text(code, str(stock.get("board") or ""))

        candidates.append(
            {
                "code": code,
                "secucode": secucode,
                "secid": secid,
                "name": str(stock.get("name") or "").strip(),
                "trade_market": market,
                "board": board,
                "industry_l3": str(stock.get("industry_l3") or "").strip(),
                "listing_date": stock.get("listing_date"),
                "is_st": "是" if is_st_stock(str(stock.get("name") or "")) else "否",
                "start_trade_date": start_bar["date"],
                "end_trade_date": end_bar["date"],
                "start_close": start_close,
                "end_close": end_close,
                "interval_change_pct": interval_change_pct,
                "interval_amplitude_pct": interval_amplitude_pct,
                "interval_trade_days": len(bars),
                "interval_high": interval_high,
                "interval_low": interval_low,
                "latest_price": stock.get("latest_price"),
                "latest_change_pct": stock.get("latest_change_pct"),
                "turnover_pct": stock.get("turnover_pct"),
                "amount": stock.get("amount"),
                "total_market_cap": stock.get("total_market_cap"),
                "float_market_cap": stock.get("float_market_cap"),
                "pe_ttm": stock.get("pe_ttm"),
                "pb": stock.get("pb"),
                "pct_60d": stock.get("pct_60d"),
                "pct_ytd": stock.get("pct_ytd"),
                "profile_source": stock.get("profile_source") or quote_source_note,
                "history_source": history_source,
            }
        )

    sorted_candidates = sort_candidates(candidates, direction=direction)
    matched_total = len(sorted_candidates)
    if top_n > 0:
        sorted_candidates = sorted_candidates[:top_n]

    detail_success_count = 0
    detail_fail_count = 0

    final_rows: list[dict[str, Any]] = []
    for row in sorted_candidates:
        used_detail = False

        if need_detail_enrichment(row):
            try:
                detail = fetch_stock_detail(client, row["secid"])
                apply_detail(row, detail)
                if detail.get("name"):
                    row["name"] = detail["name"]
                used_detail = True
                detail_success_count += 1
            except Exception:
                detail_fail_count += 1

        listing_date = row.get("listing_date")
        if not isinstance(listing_date, datetime):
            listing_date = parse_yyyymmdd(listing_date)

        listing_years = compute_listing_years(listing_date, run_time)
        data_source_parts = [str(row.get("profile_source") or ""), str(row.get("history_source") or "")]
        if used_detail:
            data_source_parts.append("stock/get@push2.eastmoney.com")
        data_source = "+".join([part for part in data_source_parts if part])

        if row["trade_market"] == "沪A":
            default_secucode = f"{row['code']}.SH"
        elif row["trade_market"] == "北A":
            default_secucode = f"{row['code']}.BJ"
        else:
            default_secucode = f"{row['code']}.SZ"

        final_rows.append(
            {
                "股票代码": row["code"],
                "证券代码": row["secucode"] or default_secucode,
                "股票简称": row["name"],
                "所属市场": row["trade_market"],
                "所属板块": row.get("board") or "",
                "所属三级行业": row.get("industry_l3") or "",
                "上市日期": to_date_text(listing_date),
                "上市年限_年": listing_years,
                "是否ST": row["is_st"],
                "区间起始交易日": to_date_text(row["start_trade_date"]),
                "区间结束交易日": to_date_text(row["end_trade_date"]),
                "区间起始收盘价": round_or_none(row.get("start_close"), 4),
                "区间结束收盘价": round_or_none(row.get("end_close"), 4),
                "区间涨跌幅_百分比": round_or_none(row.get("interval_change_pct"), 2),
                "区间振幅_百分比": round_or_none(row.get("interval_amplitude_pct"), 2),
                "区间交易天数": row.get("interval_trade_days"),
                "区间最高价": round_or_none(row.get("interval_high"), 4),
                "区间最低价": round_or_none(row.get("interval_low"), 4),
                "最新价": round_or_none(row.get("latest_price"), 4),
                "最新涨跌幅_百分比": round_or_none(row.get("latest_change_pct"), 2),
                "换手率_百分比": round_or_none(row.get("turnover_pct"), 2),
                "成交额_元": round_or_none(row.get("amount"), 2),
                "总市值_元": round_or_none(row.get("total_market_cap"), 2),
                "流通市值_元": round_or_none(row.get("float_market_cap"), 2),
                "市值区间": get_market_cap_bucket(parse_float(row.get("total_market_cap"))),
                "PE_TTM": round_or_none(row.get("pe_ttm"), 2),
                "PB": round_or_none(row.get("pb"), 2),
                "60日涨跌幅_百分比": round_or_none(row.get("pct_60d"), 2),
                "年初至今涨跌幅_百分比": round_or_none(row.get("pct_ytd"), 2),
                "数据源": data_source,
            }
        )

    history_source_display = " ; ".join([f"{key}:{val}" for key, val in history_source_counter.items()])
    if not history_source_display:
        history_source_display = "无"

    summary_rows = [
        {"指标": "执行时间", "值": run_time.strftime("%Y-%m-%d %H:%M:%S")},
        {"指标": "筛选区间", "值": f"{start_year}~{end_year}（{interval_start.strftime('%Y-%m-%d')} ~ {interval_end.strftime('%Y-%m-%d')}）"},
        {"指标": "方向", "值": direction},
        {"指标": "阈值", "值": f">= {threshold_pct}%" if direction != "fall" else f"<= -{threshold_pct}%"},
        {"指标": "TopN", "值": top_n if top_n > 0 else "不限"},
        {"指标": "MaxStocks", "值": max_stocks if max_stocks > 0 else "不限"},
        {"指标": "RequestIntervalMs", "值": args.request_interval_ms},
        {"指标": "主行情接口可用", "值": "是" if primary_quote_ok else "否"},
        {"指标": "股票池来源", "值": quote_source_note},
        {"指标": "历史行情来源", "值": history_source_display},
        {
            "指标": "主接口降级原因",
            "值": quote_fallback_reason if quote_fallback_reason else "无",
        },
        {"指标": "股票池数量", "值": universe_count},
        {"指标": "实际处理股票数", "值": processed_count},
        {"指标": "可计算区间股票数", "值": valid_interval_count},
        {"指标": "命中数量(TopN前)", "值": matched_total},
        {"指标": "输出数量", "值": len(final_rows)},
        {"指标": "历史行情失败数", "值": history_error_count},
        {"指标": "历史数据不足数", "值": insufficient_kline_count},
        {"指标": "公司画像补充成功数", "值": detail_success_count},
        {"指标": "公司画像补充失败数", "值": detail_fail_count},
        {
            "指标": "数据来源说明",
            "值": (
                "主：Eastmoney clist + push2his；"
                "降级：datacenter-web + stock/get（主接口不可用时自动切换）"
            ),
        },
    ]

    template_path = Path(__file__).resolve().parents[2] / "docs" / "templates" / "output_template.xlsx"
    if template_path.exists():
        try:
            workbook = build_workbook_with_template(template_path, summary_rows, final_rows)
        except Exception:  # noqa: BLE001
            workbook = build_workbook_fallback(summary_rows, final_rows)
    else:
        workbook = build_workbook_fallback(summary_rows, final_rows)

    workbook.save(output_path)

    payload = {
        "success": True,
        "run_time": run_time.strftime("%Y-%m-%d %H:%M:%S"),
        "start_year": start_year,
        "end_year": end_year,
        "direction": direction,
        "threshold_pct": threshold_pct,
        "universe_count": universe_count,
        "processed_count": processed_count,
        "matched_before_topn": matched_total,
        "result_count": len(final_rows),
        "quote_source": quote_source_note,
        "history_source": history_source_display,
        "output_excel": str(output_path.resolve()),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
