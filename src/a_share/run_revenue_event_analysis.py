from __future__ import annotations

import argparse
import calendar
import json
import re
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="营收同比公告事件窗口分析（Python版）")
    parser.add_argument("--codes", "--Codes", dest="codes", nargs="*", default=[], help="股票代码列表，可用逗号/空格/分号分隔")
    parser.add_argument("--listing-date-from", "--ListingDateFrom", dest="listing_date_from", default="2023-01-01")
    parser.add_argument("--window-months", "--WindowMonths", dest="window_months", type=int, default=2)
    parser.add_argument("--yoy-threshold", "--YoyThreshold", dest="yoy_threshold", type=float, default=20.0)
    parser.add_argument("--profit-threshold", "--ProfitThreshold", dest="profit_threshold", type=float, default=20.0)
    parser.add_argument("--loss-threshold", "--LossThreshold", dest="loss_threshold", type=float, default=20.0)
    parser.add_argument("--output-dir", "--OutputDir", dest="output_dir", default=r"D:\codex_1\output")
    parser.add_argument("--output-file", "--OutputFile", dest="output_file", default="")
    parser.add_argument("--json-only", "--JsonOnly", dest="json_only", action="store_true")
    parser.add_argument("--max-retry", "--MaxRetry", dest="max_retry", type=int, default=3)
    parser.add_argument("--timeout-sec", "--TimeoutSec", dest="timeout_sec", type=int, default=30)
    return parser.parse_args()


class RetryClient:
    def __init__(self, max_retry: int = 3, timeout_sec: int = 30) -> None:
        self.max_retry = max_retry
        self.timeout_sec = timeout_sec
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            }
        )

    def get_json(self, url: str, headers: dict[str, str] | None = None) -> dict[str, Any]:
        last_exc: Exception | None = None
        for attempt in range(1, self.max_retry + 1):
            try:
                resp = self.session.get(url, headers=headers, timeout=self.timeout_sec)
                resp.raise_for_status()
                return resp.json()
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if attempt >= self.max_retry:
                    break
                time.sleep(attempt)
        raise RuntimeError(f"Request failed after retries: {url}") from last_exc


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = str(value).strip()
    if not raw or raw in {"None", "null", "--", "-"}:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        pass

    if len(raw) >= 10 and re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", raw):
        try:
            return datetime.strptime(raw[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def parse_numeric(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def get_quarter_from_date(date_value: date) -> int | None:
    return {3: 1, 6: 2, 9: 3, 12: 4}.get(date_value.month)


def get_prob_pct(numerator: int, denominator: int) -> float | None:
    if denominator <= 0:
        return None
    return round((numerator * 100.0) / denominator, 2)


def get_market_prefix(secu_code: str, code: str) -> str:
    upper = (secu_code or "").upper()
    if upper.endswith(".SH"):
        return "SH"
    if upper.endswith(".SZ"):
        return "SZ"
    if upper.endswith(".BJ"):
        return "BJ"

    if code.startswith(("6", "9")):
        return "SH"
    if code.startswith(("0", "3")):
        return "SZ"
    return "BJ"


def get_secid(secu_code: str, code: str) -> str:
    upper = (secu_code or "").upper()
    if upper.endswith(".SH"):
        return f"1.{code}"
    return f"0.{code}"


def add_months(dt: date, months: int) -> date:
    month_idx = dt.month - 1 + months
    new_year = dt.year + month_idx // 12
    new_month = month_idx % 12 + 1
    max_day = calendar.monthrange(new_year, new_month)[1]
    new_day = min(dt.day, max_day)
    return date(new_year, new_month, new_day)


def first_item(value: Any) -> dict[str, Any]:
    if isinstance(value, list):
        if value and isinstance(value[0], dict):
            return value[0]
        return {}
    if isinstance(value, dict):
        return value
    return {}


def get_codes_listed_since_date(client: RetryClient, date_from: str) -> list[dict[str, Any]]:
    encoded_filter = requests.utils.quote(f"(LISTING_DATE>='{date_from}')", safe="")
    template = (
        "https://datacenter-web.eastmoney.com/api/data/v1/get"
        "?reportName=RPTA_APP_IPOAPPLY"
        "&columns=SECURITY_CODE,SECUCODE,SECURITY_NAME_ABBR,LISTING_DATE,MARKET,TRADE_MARKET"
        "&source=WEB&client=WEB&pageNumber={page}&pageSize=500"
        "&sortColumns=LISTING_DATE&sortTypes=1"
        f"&filter={encoded_filter}"
    )

    all_rows: list[dict[str, Any]] = []
    page = 1
    pages = 1
    while page <= pages:
        url = template.format(page=page)
        resp = client.get_json(url)
        if not resp.get("success") or not resp.get("result"):
            raise RuntimeError(f"Failed to fetch stock list for page {page}")
        result = resp.get("result") or {}
        pages = int(result.get("pages") or 0)
        data = result.get("data") or []
        if isinstance(data, list):
            all_rows.extend([row for row in data if isinstance(row, dict)])
        page += 1

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in all_rows:
        code = str(row.get("SECURITY_CODE") or "").strip()
        if not code:
            continue
        grouped.setdefault(code, []).append(row)

    deduped: list[dict[str, Any]] = []
    for code in sorted(grouped):
        rows = grouped[code]
        rows_sorted = sorted(
            rows,
            key=lambda x: parse_date(x.get("LISTING_DATE")) or date.max,
        )
        deduped.append(rows_sorted[0])
    return deduped


def get_company_info(client: RetryClient, code_row: dict[str, Any]) -> dict[str, Any]:
    code = str(code_row.get("SECURITY_CODE") or "").strip()
    secu_code = str(code_row.get("SECUCODE") or "").strip()
    market_prefix = get_market_prefix(secu_code, code)
    secid = get_secid(secu_code, code)

    survey_url = f"https://emweb.securities.eastmoney.com/PC_HSF10/CompanySurvey/PageAjax?code={market_prefix}{code}"
    survey = client.get_json(survey_url)
    jb = first_item((survey or {}).get("jbzl"))
    fx = first_item((survey or {}).get("fxxg"))

    quote_url = f"https://push2his.eastmoney.com/api/qt/stock/get?secid={secid}&fields=f57,f58,f116"
    quote = client.get_json(quote_url)
    quote_data = (quote or {}).get("data") or {}
    if not isinstance(quote_data, dict):
        quote_data = {}

    listing_date = parse_date(fx.get("LISTING_DATE")) or parse_date(code_row.get("LISTING_DATE"))

    industry_chain = str(jb.get("EM2016") or "").strip()
    industry_l3 = ""
    if industry_chain:
        industry_l3 = industry_chain.split("-")[-1].strip()

    name = ""
    if jb.get("SECURITY_NAME_ABBR"):
        name = str(jb.get("SECURITY_NAME_ABBR") or "")
    elif quote_data.get("f58"):
        name = str(quote_data.get("f58") or "")
    elif code_row.get("SECURITY_NAME_ABBR"):
        name = str(code_row.get("SECURITY_NAME_ABBR") or "")

    market_cap = parse_numeric(quote_data.get("f116"))
    return {
        "Code": code,
        "SecuCode": secu_code,
        "Name": name,
        "ListingDate": listing_date,
        "MarketCap": market_cap,
        "IndustryL3": industry_l3,
        "IndustryChain": industry_chain,
        "MarketPrefix": market_prefix,
        "SecId": secid,
    }


def get_financial_rows(client: RetryClient, code: str) -> list[dict[str, Any]]:
    fin_url = (
        "https://datacenter-web.eastmoney.com/api/data/v1/get"
        "?reportName=RPT_LICO_FN_CPD"
        "&columns=SECURITY_CODE,SECURITY_NAME_ABBR,REPORTDATE,NOTICE_DATE,TOTAL_OPERATE_INCOME,YSTZ,DATATYPE"
        "&source=WEB&client=WEB&pageNumber=1&pageSize=500"
        f"&filter=(SECURITY_CODE%3D%22{code}%22)"
    )
    resp = client.get_json(fin_url)
    data = ((resp or {}).get("result") or {}).get("data") or []
    if not isinstance(data, list):
        return []

    rows: list[dict[str, Any]] = []
    for raw in data:
        if not isinstance(raw, dict):
            continue
        if not raw.get("REPORTDATE") or not raw.get("NOTICE_DATE") or not raw.get("TOTAL_OPERATE_INCOME"):
            continue

        report_date = parse_date(raw.get("REPORTDATE"))
        notice_date = parse_date(raw.get("NOTICE_DATE"))
        if report_date is None or notice_date is None:
            continue

        quarter = get_quarter_from_date(report_date)
        if quarter is None:
            continue

        income = parse_numeric(raw.get("TOTAL_OPERATE_INCOME"))
        if income is None:
            continue

        rows.append(
            {
                "Year": report_date.year,
                "Quarter": quarter,
                "ReportDate": report_date,
                "NoticeDate": notice_date,
                "CumRevenue": income,
            }
        )
    return sorted(rows, key=lambda x: x["ReportDate"])


def build_single_quarter_map(financial_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    cum_map: dict[str, dict[str, Any]] = {}
    for row in financial_rows:
        key = f"{row['Year']}-Q{row['Quarter']}"
        cum_map[key] = row

    single_map: dict[str, dict[str, Any]] = {}
    for row in financial_rows:
        year = int(row["Year"])
        quarter = int(row["Quarter"])
        single_revenue: float | None = None

        if quarter == 1:
            single_revenue = float(row["CumRevenue"])
        elif quarter == 2:
            q1 = f"{year}-Q1"
            if q1 in cum_map:
                single_revenue = float(row["CumRevenue"]) - float(cum_map[q1]["CumRevenue"])
        elif quarter == 3:
            q2 = f"{year}-Q2"
            if q2 in cum_map:
                single_revenue = float(row["CumRevenue"]) - float(cum_map[q2]["CumRevenue"])
        elif quarter == 4:
            q3 = f"{year}-Q3"
            if q3 in cum_map:
                single_revenue = float(row["CumRevenue"]) - float(cum_map[q3]["CumRevenue"])

        if single_revenue is not None:
            key = f"{year}-Q{quarter}"
            single_map[key] = {
                "Year": year,
                "Quarter": quarter,
                "ReportDate": row["ReportDate"],
                "NoticeDate": row["NoticeDate"],
                "SingleRevenue": float(single_revenue),
            }
    return single_map


def get_price_rows(client: RetryClient, secid: str) -> list[dict[str, Any]]:
    k_url = (
        "https://push2his.eastmoney.com/api/qt/stock/kline/get"
        f"?secid={secid}&fields1=f1,f2,f3,f4,f5,f6"
        "&fields2=f51,f52,f53,f54,f55,f56&klt=101&fqt=1&beg=0&end=20500101"
    )
    resp = client.get_json(k_url)
    klines = ((resp or {}).get("data") or {}).get("klines") or []
    if not isinstance(klines, list):
        return []

    rows: list[dict[str, Any]] = []
    for line in klines:
        arr = str(line).split(",")
        if len(arr) < 6:
            continue
        d = parse_date(arr[0])
        close = parse_numeric(arr[2])
        high = parse_numeric(arr[3])
        low = parse_numeric(arr[4])
        if d is None or close is None or high is None or low is None:
            continue
        rows.append({"Date": d, "Close": close, "High": high, "Low": low})

    return sorted(rows, key=lambda x: x["Date"])


def analyze_one_code(
    client: RetryClient,
    code_row: dict[str, Any],
    window_months: int,
    yoy_threshold: float,
    profit_threshold: float,
    loss_threshold: float,
) -> dict[str, Any]:
    code = str(code_row.get("SECURITY_CODE") or "")
    error_message = ""

    try:
        info = get_company_info(client, code_row)
    except Exception as exc:  # noqa: BLE001
        error_message = f"CompanyInfoError: {exc}"
        fallback_secu_code = str(code_row.get("SECUCODE") or "")
        info = {
            "Code": code,
            "SecuCode": fallback_secu_code,
            "Name": str(code_row.get("SECURITY_NAME_ABBR") or ""),
            "ListingDate": parse_date(code_row.get("LISTING_DATE")),
            "MarketCap": None,
            "IndustryL3": "",
            "IndustryChain": "",
            "MarketPrefix": get_market_prefix(fallback_secu_code, code),
            "SecId": get_secid(fallback_secu_code, code),
        }

    event_rows: list[dict[str, Any]] = []
    stats: dict[str, Any] = {
        "SignalCountAll": 0,
        "GainHitCountAll": 0,
        "GainHitProbAllPct": None,
        "LossHitCountAll": 0,
        "LossHitProbAllPct": None,
        "SignalCountCompleted": 0,
        "GainHitCountCompleted": 0,
        "GainHitProbCompletedPct": None,
        "LossHitCountCompleted": 0,
        "LossHitProbCompletedPct": None,
    }

    try:
        fin_rows = get_financial_rows(client, code)
        if fin_rows:
            single_map = build_single_quarter_map(fin_rows)
            price_rows = get_price_rows(client, str(info["SecId"]))
            if price_rows and single_map:
                latest_price_date = price_rows[-1]["Date"]
                for key in sorted(single_map.keys()):
                    curr = single_map[key]
                    prev_key = f"{curr['Year'] - 1}-Q{curr['Quarter']}"
                    if prev_key not in single_map:
                        continue

                    prev = single_map[prev_key]
                    prev_single = float(prev["SingleRevenue"])
                    if prev_single == 0:
                        continue

                    yoy = ((float(curr["SingleRevenue"]) - prev_single) / prev_single) * 100.0
                    if yoy <= yoy_threshold:
                        continue

                    start_trade = next((p for p in price_rows if p["Date"] >= curr["NoticeDate"]), None)
                    if start_trade is None:
                        event_rows.append(
                            {
                                "Quarter": f"{curr['Year']}Q{curr['Quarter']}",
                                "NoticeDate": curr["NoticeDate"],
                                "SingleYoyPct": round(yoy, 2),
                                "BaseDate": None,
                                "BaseClose": None,
                                "WindowEnd": add_months(curr["NoticeDate"], window_months),
                                "CompletedWindow": False,
                                "MaxUpPct": None,
                                "MaxUpDate": None,
                                "MaxDownPct": None,
                                "MaxDownDate": None,
                                "GainHit": False,
                                "LossHit": False,
                                "Note": "公告后无可用交易日",
                            }
                        )
                        continue

                    window_end = add_months(curr["NoticeDate"], window_months)
                    window_prices = [p for p in price_rows if start_trade["Date"] <= p["Date"] <= window_end]
                    if not window_prices:
                        event_rows.append(
                            {
                                "Quarter": f"{curr['Year']}Q{curr['Quarter']}",
                                "NoticeDate": curr["NoticeDate"],
                                "SingleYoyPct": round(yoy, 2),
                                "BaseDate": start_trade["Date"],
                                "BaseClose": round(float(start_trade["Close"]), 3),
                                "WindowEnd": window_end,
                                "CompletedWindow": latest_price_date >= window_end,
                                "MaxUpPct": None,
                                "MaxUpDate": None,
                                "MaxDownPct": None,
                                "MaxDownDate": None,
                                "GainHit": False,
                                "LossHit": False,
                                "Note": "窗口内无行情数据",
                            }
                        )
                        continue

                    base_price = float(start_trade["Close"])
                    max_high = max(window_prices, key=lambda x: x["High"])
                    min_low = min(window_prices, key=lambda x: x["Low"])
                    max_up_pct = ((float(max_high["High"]) - base_price) / base_price) * 100.0
                    max_down_pct = ((float(min_low["Low"]) - base_price) / base_price) * 100.0
                    completed = latest_price_date >= window_end

                    event_rows.append(
                        {
                            "Quarter": f"{curr['Year']}Q{curr['Quarter']}",
                            "NoticeDate": curr["NoticeDate"],
                            "SingleYoyPct": round(yoy, 2),
                            "BaseDate": start_trade["Date"],
                            "BaseClose": round(base_price, 3),
                            "WindowEnd": window_end,
                            "CompletedWindow": completed,
                            "MaxUpPct": round(max_up_pct, 2),
                            "MaxUpDate": max_high["Date"],
                            "MaxDownPct": round(max_down_pct, 2),
                            "MaxDownDate": min_low["Date"],
                            "GainHit": max_up_pct > profit_threshold,
                            "LossHit": max_down_pct < (-1.0 * loss_threshold),
                            "Note": "",
                        }
                    )

        valid_events = [x for x in event_rows if x["MaxUpPct"] is not None and x["MaxDownPct"] is not None]
        completed_events = [x for x in valid_events if bool(x["CompletedWindow"])]

        stats["SignalCountAll"] = len(valid_events)
        stats["GainHitCountAll"] = sum(1 for x in valid_events if bool(x["GainHit"]))
        stats["LossHitCountAll"] = sum(1 for x in valid_events if bool(x["LossHit"]))
        stats["GainHitProbAllPct"] = get_prob_pct(stats["GainHitCountAll"], stats["SignalCountAll"])
        stats["LossHitProbAllPct"] = get_prob_pct(stats["LossHitCountAll"], stats["SignalCountAll"])

        stats["SignalCountCompleted"] = len(completed_events)
        stats["GainHitCountCompleted"] = sum(1 for x in completed_events if bool(x["GainHit"]))
        stats["LossHitCountCompleted"] = sum(1 for x in completed_events if bool(x["LossHit"]))
        stats["GainHitProbCompletedPct"] = get_prob_pct(
            stats["GainHitCountCompleted"], stats["SignalCountCompleted"]
        )
        stats["LossHitProbCompletedPct"] = get_prob_pct(
            stats["LossHitCountCompleted"], stats["SignalCountCompleted"]
        )
    except Exception as exc:  # noqa: BLE001
        if error_message:
            error_message = f"{error_message} | AnalysisError: {exc}"
        else:
            error_message = f"AnalysisError: {exc}"

    event_rows = sorted(event_rows, key=lambda x: str(x["Quarter"]))
    return {
        "Code": info["Code"],
        "Name": info["Name"],
        "SecuCode": info["SecuCode"],
        "ListingDate": info["ListingDate"],
        "MarketCap": info["MarketCap"],
        "IndustryL3": info["IndustryL3"],
        "IndustryChain": info["IndustryChain"],
        "EventRows": event_rows,
        "Stats": stats,
        "Error": error_message,
    }


def build_code_rows_from_input(input_codes: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for token in input_codes:
        if not token:
            continue
        parts = [p for p in re.split(r"[,;\s]+", token) if p]
        for part in parts:
            clean = re.sub(r"\s+", "", part)
            if not clean:
                continue
            if clean.startswith(("6", "9")):
                suffix = ".SH"
            elif clean.startswith(("0", "3")):
                suffix = ".SZ"
            else:
                suffix = ".BJ"
            rows.append(
                {
                    "SECURITY_CODE": clean,
                    "SECUCODE": f"{clean}{suffix}",
                    "SECURITY_NAME_ABBR": "",
                    "LISTING_DATE": None,
                }
            )

    dedup: dict[str, dict[str, Any]] = {}
    for row in rows:
        code = str(row.get("SECURITY_CODE") or "")
        if code and code not in dedup:
            dedup[code] = row
    return [dedup[k] for k in sorted(dedup.keys())]


def to_date_text(dt: date | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d")


def auto_fit_worksheet(ws: Any) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)


def write_rows(ws: Any, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    if not rows:
        ws.append(["" for _ in headers])
    auto_fit_worksheet(ws)


def add_table(ws: Any, table_name: str) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    client = RetryClient(max_retry=max(1, int(args.max_retry)), timeout_sec=max(5, int(args.timeout_sec)))

    if args.codes:
        code_rows = build_code_rows_from_input(args.codes)
    else:
        code_rows = get_codes_listed_since_date(client, args.listing_date_from)

    if not code_rows:
        raise RuntimeError("No stock codes to analyze.")

    output_file = args.output_file.strip() if args.output_file else ""
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"revenue_yoy_event_scan_{timestamp}.xlsx"

    excel_path = output_dir / output_file
    if excel_path.exists() and not args.json_only:
        excel_path.unlink()

    summary_rows: list[dict[str, Any]] = []
    all_results: list[dict[str, Any]] = []

    total = len(code_rows)
    for idx, row in enumerate(code_rows, start=1):
        code = str(row.get("SECURITY_CODE") or "")
        print(f"[{idx}/{total}] analyzing {code}")

        analysis = analyze_one_code(
            client=client,
            code_row=row,
            window_months=args.window_months,
            yoy_threshold=args.yoy_threshold,
            profit_threshold=args.profit_threshold,
            loss_threshold=args.loss_threshold,
        )
        all_results.append(analysis)

        summary_rows.append(
            {
                "Code": analysis["Code"],
                "Name": analysis["Name"],
                "ListingDate": to_date_text(analysis["ListingDate"]),
                "MarketCap": round(float(analysis["MarketCap"]), 2) if analysis["MarketCap"] is not None else None,
                "IndustryL3": analysis["IndustryL3"],
                "IndustryChain": analysis["IndustryChain"],
                "WindowMonths": args.window_months,
                "YoyThresholdPct": args.yoy_threshold,
                "ProfitThresholdPct": args.profit_threshold,
                "LossThresholdPct": args.loss_threshold,
                "SignalCountAll": analysis["Stats"]["SignalCountAll"],
                "GainHitCountAll": analysis["Stats"]["GainHitCountAll"],
                "GainHitProbAllPct": analysis["Stats"]["GainHitProbAllPct"],
                "LossHitCountAll": analysis["Stats"]["LossHitCountAll"],
                "LossHitProbAllPct": analysis["Stats"]["LossHitProbAllPct"],
                "SignalCountCompleted": analysis["Stats"]["SignalCountCompleted"],
                "GainHitCountCompleted": analysis["Stats"]["GainHitCountCompleted"],
                "GainHitProbCompletedPct": analysis["Stats"]["GainHitProbCompletedPct"],
                "LossHitCountCompleted": analysis["Stats"]["LossHitCountCompleted"],
                "LossHitProbCompletedPct": analysis["Stats"]["LossHitProbCompletedPct"],
                "Error": analysis["Error"],
            }
        )

    if args.json_only:
        payload = {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "parameters": {
                "codeCount": total,
                "listingDateFrom": args.listing_date_from,
                "windowMonths": args.window_months,
                "yoyThreshold": args.yoy_threshold,
                "profitThreshold": args.profit_threshold,
                "lossThreshold": args.loss_threshold,
            },
            "summary": summary_rows,
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    event_headers = [
        "Quarter",
        "NoticeDate",
        "SingleYoyPct",
        "BaseDate",
        "BaseClose",
        "WindowEnd",
        "CompletedWindow",
        "MaxUpPct",
        "MaxUpDate",
        "MaxDownPct",
        "MaxDownDate",
        "GainHit",
        "LossHit",
        "Note",
        "Error",
    ]

    for result in all_results:
        sheet_name = f"C_{result['Code']}"
        ws = workbook.create_sheet(title=sheet_name[:31])
        rows: list[dict[str, Any]] = []

        if result["EventRows"]:
            for ev in result["EventRows"]:
                rows.append(
                    {
                        "Quarter": ev["Quarter"],
                        "NoticeDate": to_date_text(ev["NoticeDate"]),
                        "SingleYoyPct": ev["SingleYoyPct"],
                        "BaseDate": to_date_text(ev["BaseDate"]),
                        "BaseClose": ev["BaseClose"],
                        "WindowEnd": to_date_text(ev["WindowEnd"]),
                        "CompletedWindow": ev["CompletedWindow"],
                        "MaxUpPct": ev["MaxUpPct"],
                        "MaxUpDate": to_date_text(ev["MaxUpDate"]),
                        "MaxDownPct": ev["MaxDownPct"],
                        "MaxDownDate": to_date_text(ev["MaxDownDate"]),
                        "GainHit": ev["GainHit"],
                        "LossHit": ev["LossHit"],
                        "Note": ev["Note"],
                        "Error": result["Error"],
                    }
                )
        else:
            rows.append(
                {
                    "Quarter": "",
                    "NoticeDate": "",
                    "SingleYoyPct": None,
                    "BaseDate": "",
                    "BaseClose": None,
                    "WindowEnd": "",
                    "CompletedWindow": None,
                    "MaxUpPct": None,
                    "MaxUpDate": "",
                    "MaxDownPct": None,
                    "MaxDownDate": "",
                    "GainHit": None,
                    "LossHit": None,
                    "Note": "无触发事件或数据不足",
                    "Error": result["Error"],
                }
            )

        write_rows(ws, rows, event_headers)
        add_table(ws, f"T_{result['Code']}"[:255])

    ws_summary = workbook.create_sheet(title="Summary")
    summary_headers = list(summary_rows[0].keys()) if summary_rows else []
    write_rows(ws_summary, summary_rows, summary_headers)
    add_table(ws_summary, "T_Summary")

    workbook.save(excel_path)

    success_count = sum(1 for x in summary_rows if not x.get("Error"))
    error_count = sum(1 for x in summary_rows if x.get("Error"))

    payload = {
        "output": str(excel_path.resolve()),
        "codeCount": total,
        "successCount": success_count,
        "errorCount": error_count,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
